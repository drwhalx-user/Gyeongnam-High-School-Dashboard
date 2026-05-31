"""
학교폭력 피해 위험 점수(school_violence_risk_score)를 산출한다.

2023~2025학년도 3개년 학교폭력 실태조사 결과를 이용해 아래 변수를 생성한다.
  - victim_rate                    : 연도별 피해율 (피해학생수 / 설문참여자수)
  - avg_victim_rate_3yr            : 3개년 평균 피해율
  - school_violence_risk_score     : avg_victim_rate_3yr의 Min-Max 정규화 (0~1)
  - violence_years_available       : 유효 연도 수 (1~3)
  - avg_survey_participation_rate_3yr : 3개년 평균 설문 참여율

이번 단계에서는 CDI 최종 점수는 산출하지 않는다.

입력 파일:
  data/raw/school_violence_survey_2023_2025.xlsx
  data/processed/gyeongnam_high_schools_master_table.xlsx (student_count 보충용)

출력 파일:
  data/processed/gyeongnam_high_schools_master_table.xlsx   (master_table 시트 갱신)
  data/processed/gyeongnam_high_schools_violence_risk.xlsx  (신규)
  outputs/tables/violence_risk_summary_2025.csv
  outputs/tables/violence_risk_missing_check_2025.csv
  docs/violence_risk_methodology_2025.md
"""

import sys
import pathlib

import numpy as np
import pandas as pd
from openpyxl import load_workbook

# ── 경로 설정 ─────────────────────────────────────────────────────────────────
ROOT          = pathlib.Path(__file__).resolve().parent.parent
RAW_DIR       = ROOT / "data" / "raw"
PROCESSED_DIR = ROOT / "data" / "processed"
TABLES_DIR    = ROOT / "outputs" / "tables"
DOCS_DIR      = ROOT / "docs"

for d in [PROCESSED_DIR, TABLES_DIR, DOCS_DIR]:
    d.mkdir(parents=True, exist_ok=True)

# ════════════════════════════════════════════════════════════════════════════
# STEP 1. 입력 파일 로드
# ════════════════════════════════════════════════════════════════════════════
print("[STEP 1] 입력 파일 로드")

VIOLENCE_PATH = RAW_DIR / "school_violence_survey_2023_2025.xlsx"
MASTER_PATH   = PROCESSED_DIR / "gyeongnam_high_schools_master_table.xlsx"

for p in [VIOLENCE_PATH, MASTER_PATH]:
    if not p.exists():
        print(f"[ERROR] 파일 없음: {p}")
        sys.exit(1)

violence_df = pd.read_excel(VIOLENCE_PATH)
print(f"  학교폭력 데이터: {violence_df.shape[0]}행 × {violence_df.shape[1]}열")
print(f"  컬럼: {list(violence_df.columns)}")

xl = pd.ExcelFile(MASTER_PATH)
sheet_used = "master_table" if "master_table" in xl.sheet_names else xl.sheet_names[0]
master_df = pd.read_excel(MASTER_PATH, sheet_name=sheet_used)
print(f"  마스터 테이블: {master_df.shape[0]}행 × {master_df.shape[1]}열 (시트: {sheet_used})")

# ════════════════════════════════════════════════════════════════════════════
# STEP 2. 2025년 total_student_count 보충 (master_table의 student_count 활용)
# ════════════════════════════════════════════════════════════════════════════
print("\n[STEP 2] 2025년 total_student_count 보충")

# master_table에서 school_code → student_count 매핑
if "student_count" not in master_df.columns:
    print("[WARN] master_table에 student_count 없음 — 보충 건너뜀")
    supp_map = {}
else:
    supp_map = master_df.set_index("school_code")["student_count"].to_dict()

mask_2025_missing = (
    (violence_df["year"] == 2025)
    & violence_df["total_student_count"].isna()
)
n_before = mask_2025_missing.sum()

if supp_map:
    violence_df.loc[mask_2025_missing, "total_student_count"] = (
        violence_df.loc[mask_2025_missing, "school_code"].map(supp_map)
    )

n_after = (
    (violence_df["year"] == 2025) & violence_df["total_student_count"].isna()
).sum()
print(f"  2025년 결측 total_student_count: {n_before}개 → 보충 후 {n_after}개")

# ════════════════════════════════════════════════════════════════════════════
# STEP 3. 연도별 victim_rate 및 survey_participation_rate 계산
# ════════════════════════════════════════════════════════════════════════════
print("\n[STEP 3] 연도별 피해율 및 참여율 계산")

violence_df["victim_rate"] = (
    violence_df["victim_student_count"] / violence_df["survey_participant_count"]
)
violence_df["survey_participation_rate"] = (
    violence_df["survey_participant_count"] / violence_df["total_student_count"]
)

print(f"  victim_rate 결측: {violence_df['victim_rate'].isna().sum()}개")
print(f"  victim_rate 범위: {violence_df['victim_rate'].min():.4f} ~ {violence_df['victim_rate'].max():.4f}")
print(f"  survey_participation_rate 결측: {violence_df['survey_participation_rate'].isna().sum()}개")

# ════════════════════════════════════════════════════════════════════════════
# STEP 4. 학교별 3개년 평균 집계
# ════════════════════════════════════════════════════════════════════════════
print("\n[STEP 4] 학교별 3개년 평균 집계")

agg = (
    violence_df
    .groupby("school_code")
    .agg(
        school_name          = ("school_name", "first"),
        sigungu              = ("sigungu",      "first"),
        avg_victim_rate_3yr  = ("victim_rate",               "mean"),
        violence_years_available = ("victim_rate",           lambda x: x.notna().sum()),
        avg_survey_participation_rate_3yr = ("survey_participation_rate", "mean"),
    )
    .reset_index()
)

print(f"  집계 후 학교 수: {len(agg)}개교")
print(f"  avg_victim_rate_3yr 결측: {agg['avg_victim_rate_3yr'].isna().sum()}개교")
print(f"  violence_years_available 분포:\n{agg['violence_years_available'].value_counts().sort_index().to_string()}")

# ════════════════════════════════════════════════════════════════════════════
# STEP 5. Min-Max 정규화 → school_violence_risk_score
# ════════════════════════════════════════════════════════════════════════════
print("\n[STEP 5] Min-Max 정규화 → school_violence_risk_score")

valid = agg["avg_victim_rate_3yr"].dropna()
vmin, vmax = valid.min(), valid.max()
print(f"  avg_victim_rate_3yr: min={vmin:.6f}, max={vmax:.6f}")

if vmax == vmin:
    # 모든 값이 동일한 경우
    agg["school_violence_risk_score"] = agg["avg_victim_rate_3yr"].apply(
        lambda x: 0.5 if pd.notna(x) else float("nan")
    )
    print("  [WARN] min == max → 모든 학교 0.5 부여")
else:
    agg["school_violence_risk_score"] = (
        (agg["avg_victim_rate_3yr"] - vmin) / (vmax - vmin)
    )

print(f"  school_violence_risk_score 결측: {agg['school_violence_risk_score'].isna().sum()}개교")
print(f"  school_violence_risk_score 범위: {agg['school_violence_risk_score'].min():.4f} ~ {agg['school_violence_risk_score'].max():.4f}")
print(f"  school_violence_risk_score 평균: {agg['school_violence_risk_score'].mean():.4f}")

# ════════════════════════════════════════════════════════════════════════════
# STEP 6. 마스터 테이블에 변수 병합
# ════════════════════════════════════════════════════════════════════════════
print("\n[STEP 6] 마스터 테이블에 변수 병합")

MERGE_COLS = [
    "school_code",
    "avg_victim_rate_3yr",
    "school_violence_risk_score",
    "violence_years_available",
    "avg_survey_participation_rate_3yr",
]

# 기존 컬럼 중복 제거
for col in MERGE_COLS[1:]:
    if col in master_df.columns:
        master_df.drop(columns=[col], inplace=True)
        print(f"  기존 컬럼 제거: {col}")

master_updated = master_df.merge(
    agg[MERGE_COLS],
    on="school_code",
    how="left",
)

n_matched = master_updated["school_violence_risk_score"].notna().sum()
print(f"  병합 결과: {len(master_updated)}행, school_violence_risk_score 비결측={n_matched}개교")

# ════════════════════════════════════════════════════════════════════════════
# STEP 7. master_table.xlsx 저장 (기존 시트 보존, openpyxl 활용)
# ════════════════════════════════════════════════════════════════════════════
print("\n[STEP 7] master_table.xlsx 저장")

# 요약 / 결측 점검 시트 생성
summary_cols = [
    "school_code", "school_name", "sigungu",
    "avg_victim_rate_3yr", "school_violence_risk_score",
    "violence_years_available", "avg_survey_participation_rate_3yr",
]
summary_df = master_updated[[c for c in summary_cols if c in master_updated.columns]].copy()

missing_check_df = summary_df[summary_df["school_violence_risk_score"].isna()].copy()

wb = load_workbook(MASTER_PATH)

# master_table 시트 교체
if "master_table" in wb.sheetnames:
    del wb["master_table"]
ws_master = wb.create_sheet("master_table", 0)

# 데이터 쓰기 — 헤더
for c_idx, col_name in enumerate(master_updated.columns, start=1):
    ws_master.cell(row=1, column=c_idx, value=col_name)

# 데이터 행
for r_idx, row in enumerate(master_updated.itertuples(index=False), start=2):
    for c_idx, val in enumerate(row, start=1):
        if isinstance(val, float) and np.isnan(val):
            ws_master.cell(row=r_idx, column=c_idx, value=None)
        else:
            ws_master.cell(row=r_idx, column=c_idx, value=val)

# violence_risk_summary 시트 추가
SHEET_SUMMARY = "violence_risk_summary"
if SHEET_SUMMARY in wb.sheetnames:
    del wb[SHEET_SUMMARY]
ws_sum = wb.create_sheet(SHEET_SUMMARY)
for c_idx, col_name in enumerate(summary_df.columns, start=1):
    ws_sum.cell(row=1, column=c_idx, value=col_name)
for r_idx, row in enumerate(summary_df.itertuples(index=False), start=2):
    for c_idx, val in enumerate(row, start=1):
        if isinstance(val, float) and np.isnan(val):
            ws_sum.cell(row=r_idx, column=c_idx, value=None)
        else:
            ws_sum.cell(row=r_idx, column=c_idx, value=val)

# violence_risk_missing 시트 추가
SHEET_MISSING = "violence_risk_missing"
if SHEET_MISSING in wb.sheetnames:
    del wb[SHEET_MISSING]
ws_mis = wb.create_sheet(SHEET_MISSING)
if len(missing_check_df) > 0:
    for c_idx, col_name in enumerate(missing_check_df.columns, start=1):
        ws_mis.cell(row=1, column=c_idx, value=col_name)
    for r_idx, row in enumerate(missing_check_df.itertuples(index=False), start=2):
        for c_idx, val in enumerate(row, start=1):
            if isinstance(val, float) and np.isnan(val):
                ws_mis.cell(row=r_idx, column=c_idx, value=None)
            else:
                ws_mis.cell(row=r_idx, column=c_idx, value=val)
else:
    ws_mis.cell(row=1, column=1, value="결측 없음")

wb.save(MASTER_PATH)
print(f"  저장 완료: {MASTER_PATH}")
print(f"  시트 목록: {wb.sheetnames}")

# ════════════════════════════════════════════════════════════════════════════
# STEP 8. violence_risk.xlsx 저장
# ════════════════════════════════════════════════════════════════════════════
print("\n[STEP 8] violence_risk.xlsx 저장")

VIOLENCE_RISK_PATH = PROCESSED_DIR / "gyeongnam_high_schools_violence_risk.xlsx"

with pd.ExcelWriter(VIOLENCE_RISK_PATH, engine="openpyxl") as writer:
    summary_df.to_excel(writer, sheet_name="violence_risk_summary", index=False)
    missing_check_df.to_excel(writer, sheet_name="violence_risk_missing",  index=False)

    # 연도별 상세 데이터 시트
    detail_cols = [
        "school_code", "school_name", "sigungu", "year",
        "victim_student_count", "survey_participant_count",
        "total_student_count", "victim_rate", "survey_participation_rate",
    ]
    violence_df[[c for c in detail_cols if c in violence_df.columns]].sort_values(
        ["sigungu", "school_name", "year"]
    ).to_excel(writer, sheet_name="yearly_detail", index=False)

print(f"  저장 완료: {VIOLENCE_RISK_PATH}")

# ════════════════════════════════════════════════════════════════════════════
# STEP 9. 요약 CSV 및 결측 점검 CSV 저장
# ════════════════════════════════════════════════════════════════════════════
print("\n[STEP 9] CSV 저장")

SUMMARY_CSV = TABLES_DIR / "violence_risk_summary_2025.csv"
MISSING_CSV = TABLES_DIR / "violence_risk_missing_check_2025.csv"

summary_df.to_csv(SUMMARY_CSV,      index=False, encoding="utf-8-sig")
missing_check_df.to_csv(MISSING_CSV, index=False, encoding="utf-8-sig")

print(f"  요약 CSV:    {SUMMARY_CSV}")
print(f"  결측 점검 CSV: {MISSING_CSV}")

# ════════════════════════════════════════════════════════════════════════════
# STEP 10. 방법론 문서 저장
# ════════════════════════════════════════════════════════════════════════════
print("\n[STEP 10] 방법론 문서 저장")

methodology_md = f"""# 학교폭력 피해 위험 점수 산출 방법론

## 1. 개요
2023~2025학년도 학교폭력 실태조사 결과를 바탕으로 경남 일반고 146개교의
학교폭력 피해 위험 점수를 산출한다.

## 2. 입력 데이터
| 파일 | 설명 |
|------|------|
| `data/raw/school_violence_survey_2023_2025.xlsx` | 연도별 학교폭력 실태조사 결과 (3개년 438행) |
| `data/processed/gyeongnam_high_schools_master_table.xlsx` | 2025년 total_student_count 보충용 |

## 3. 변수 정의

### 3.1 연도별 계산
| 변수 | 산식 |
|------|------|
| `victim_rate` | 피해학생수 ÷ 설문참여자수 |
| `survey_participation_rate` | 설문참여자수 ÷ 전체학생수 |

### 3.2 3개년 집계
| 변수 | 산식 |
|------|------|
| `avg_victim_rate_3yr` | 3개년 victim_rate 산술평균 (결측 연도 제외) |
| `violence_years_available` | 유효 연도 수 (1~3) |
| `avg_survey_participation_rate_3yr` | 3개년 survey_participation_rate 산술평균 |

### 3.3 정규화
| 변수 | 산식 |
|------|------|
| `school_violence_risk_score` | (avg_victim_rate_3yr − min) ÷ (max − min) |

- **정규화 기준**: 경남 일반고 146개교 전체
- **해석**: 0에 가까울수록 피해율 낮음, 1에 가까울수록 피해율 높음

## 4. 결측 처리
- 2025년 total_student_count 결측: master_table의 student_count로 보충
- victim_student_count 또는 survey_participant_count 결측 → victim_rate = NaN
- 유효 연도가 0인 학교 → avg_victim_rate_3yr = NaN, school_violence_risk_score = NaN

## 5. 산출 결과 요약
- 대상 학교: {len(agg)}개교
- school_violence_risk_score 비결측: {agg['school_violence_risk_score'].notna().sum()}개교
- avg_victim_rate_3yr: 최솟값={vmin:.6f}, 최댓값={vmax:.6f}
- school_violence_risk_score: 평균={agg['school_violence_risk_score'].mean():.4f}

## 6. 출력 파일
| 파일 | 설명 |
|------|------|
| `data/processed/gyeongnam_high_schools_master_table.xlsx` | master_table 시트 갱신 |
| `data/processed/gyeongnam_high_schools_violence_risk.xlsx` | 위험 점수 요약 및 연도별 상세 |
| `outputs/tables/violence_risk_summary_2025.csv` | 학교별 요약 |
| `outputs/tables/violence_risk_missing_check_2025.csv` | 결측 학교 목록 |

## 7. 주의 사항
- 이 단계에서는 CDI 최종 점수를 산출하지 않는다.
- CDI = (demand_size_score + counseling_use_score + school_violence_risk_score) / 3 는 별도 스크립트에서 산출 예정.
"""

METHODOLOGY_MD = DOCS_DIR / "violence_risk_methodology_2025.md"
METHODOLOGY_MD.write_text(methodology_md, encoding="utf-8")
print(f"  방법론 문서: {METHODOLOGY_MD}")

# ════════════════════════════════════════════════════════════════════════════
# 최종 요약 출력
# ════════════════════════════════════════════════════════════════════════════
print("\n" + "=" * 60)
print("학교폭력 피해 위험 점수 산출 완료")
print("=" * 60)
print(f"\n대상 학교: {len(agg)}개교")
print(f"violence_years_available 분포:")
for yr, cnt in agg["violence_years_available"].value_counts().sort_index().items():
    print(f"  {yr}개년: {cnt}개교")
print(f"\navg_victim_rate_3yr 결측: {agg['avg_victim_rate_3yr'].isna().sum()}개교")
print(f"school_violence_risk_score 결측: {agg['school_violence_risk_score'].isna().sum()}개교")
print(f"school_violence_risk_score 평균: {agg['school_violence_risk_score'].mean():.4f}")
print(f"avg_survey_participation_rate_3yr 평균: {agg['avg_survey_participation_rate_3yr'].mean():.4f}")
print("\n생성 파일:")
print(f"  {MASTER_PATH}")
print(f"  {VIOLENCE_RISK_PATH}")
print(f"  {SUMMARY_CSV}")
print(f"  {MISSING_CSV}")
print(f"  {METHODOLOGY_MD}")
