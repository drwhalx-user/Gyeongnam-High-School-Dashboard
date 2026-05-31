"""
상담수요지수(CDI) 산출, 등급화 및 master table 병합

CDI = (demand_size_score + counseling_use_score + school_violence_risk_score) / 3

이번 단계에서는 Priority Score를 계산하지 않는다.

입력 파일:
  data/processed/gyeongnam_high_schools_master_table.xlsx

출력 파일:
  data/processed/gyeongnam_high_schools_master_table.xlsx   (master_table 시트 갱신)
  data/processed/gyeongnam_high_schools_CDI.xlsx            (신규)
  outputs/tables/cdi_summary_2025.csv
  outputs/tables/cdi_missing_check_2025.csv
  docs/cdi_methodology_2025.md
"""

import sys
import pathlib

import numpy as np
import pandas as pd
from openpyxl import load_workbook

# ── 경로 설정 ─────────────────────────────────────────────────────────────────
ROOT          = pathlib.Path(__file__).resolve().parent.parent
PROCESSED_DIR = ROOT / "data" / "processed"
TABLES_DIR    = ROOT / "outputs" / "tables"
DOCS_DIR      = ROOT / "docs"

for d in [PROCESSED_DIR, TABLES_DIR, DOCS_DIR]:
    d.mkdir(parents=True, exist_ok=True)

MASTER_PATH = PROCESSED_DIR / "gyeongnam_high_schools_master_table.xlsx"
CDI_PATH    = PROCESSED_DIR / "gyeongnam_high_schools_CDI.xlsx"

# ════════════════════════════════════════════════════════════════════════════
# STEP 1. 입력 파일 로드
# ════════════════════════════════════════════════════════════════════════════
print("[STEP 1] 입력 파일 로드")

if not MASTER_PATH.exists():
    print(f"[ERROR] 파일 없음: {MASTER_PATH}")
    sys.exit(1)

xl = pd.ExcelFile(MASTER_PATH)
if "master_table" in xl.sheet_names:
    sheet_used = "master_table"
else:
    sheet_used = xl.sheet_names[0]
    print(f"  [WARN] 'master_table' 시트 없음 — '{sheet_used}' 시트 사용")

master_df = pd.read_excel(MASTER_PATH, sheet_name=sheet_used)
n_rows_original = len(master_df)
print(f"  로드 완료: {n_rows_original}행 × {len(master_df.columns)}열 (시트: {sheet_used})")

# ════════════════════════════════════════════════════════════════════════════
# STEP 2. 필수 변수 존재 확인
# ════════════════════════════════════════════════════════════════════════════
print("\n[STEP 2] 필수 변수 확인")

ID_COLS  = ["school_code", "school_name", "sido", "sigungu"]
CDI_COLS = ["demand_size_score", "counseling_use_score", "school_violence_risk_score"]
ALL_REQUIRED = ID_COLS + CDI_COLS

missing_vars = [c for c in ALL_REQUIRED if c not in master_df.columns]
if missing_vars:
    # 유사 변수명 탐색
    print(f"  [ERROR] 필수 변수 없음: {missing_vars}")
    print("  유사 변수명 탐색:")
    for mv in missing_vars:
        candidates = [c for c in master_df.columns if mv[:6].lower() in c.lower()]
        print(f"    '{mv}' 후보: {candidates if candidates else '없음'}")
    sys.exit(1)

print("  필수 변수 전체 확인 완료")
for col in CDI_COLS:
    n_valid = master_df[col].notna().sum()
    print(f"    {col}: {n_valid}/{n_rows_original} 비결측")

# ════════════════════════════════════════════════════════════════════════════
# STEP 3. CDI 계산
# ════════════════════════════════════════════════════════════════════════════
print("\n[STEP 3] CDI 계산")

df = master_df.copy()

# 세 구성요소가 모두 존재할 때만 공식 CDI 계산
all_available_mask = df[CDI_COLS].notna().all(axis=1)
df["CDI"] = np.where(
    all_available_mask,
    df[CDI_COLS].mean(axis=1),
    np.nan,
)

# 결측 제외 평균 (민감도 분석용)
df["cdi_available_mean"] = df[CDI_COLS].mean(axis=1, skipna=True)

# 비결측 구성요소 수
df["cdi_components_available"] = df[CDI_COLS].notna().sum(axis=1).astype(int)

# 결측 구성요소명 목록 (문자열)
def get_missing_components(row):
    missing = [col for col in CDI_COLS if pd.isna(row[col])]
    return ", ".join(missing) if missing else "없음"

df["cdi_missing_components"] = df[CDI_COLS].apply(get_missing_components, axis=1)

n_cdi_valid = df["CDI"].notna().sum()
n_cdi_missing = df["CDI"].isna().sum()
print(f"  CDI 산출: {n_cdi_valid}개교 / 결측: {n_cdi_missing}개교")
if n_cdi_valid > 0:
    print(f"  CDI 범위: {df['CDI'].min():.4f} ~ {df['CDI'].max():.4f}")
    print(f"  CDI 평균: {df['CDI'].mean():.4f}")

# ════════════════════════════════════════════════════════════════════════════
# STEP 4. CDI 등급화 및 그룹화
# ════════════════════════════════════════════════════════════════════════════
print("\n[STEP 4] CDI 등급화 및 분위수 그룹화")

# 4-1. 절대값 기준 cdi_level
def assign_cdi_level(cdi):
    if pd.isna(cdi):
        return "확인 필요"
    elif cdi < 0.3:
        return "수요 낮음"
    elif cdi < 0.6:
        return "수요 보통"
    elif cdi < 0.8:
        return "수요 높음"
    else:
        return "수요 매우 높음"

df["cdi_level"] = df["CDI"].apply(assign_cdi_level)

level_order = ["수요 낮음", "수요 보통", "수요 높음", "수요 매우 높음", "확인 필요"]
level_counts = df["cdi_level"].value_counts()
print("  cdi_level 분포:")
for lv in level_order:
    cnt = level_counts.get(lv, 0)
    pct = cnt / n_rows_original * 100
    print(f"    {lv}: {cnt}개교 ({pct:.1f}%)")

# 4-2. 분위수 기준 cdi_quantile_group (결측 제외)
valid_cdi = df.loc[df["CDI"].notna(), "CDI"]
p20 = valid_cdi.quantile(0.20)
p80 = valid_cdi.quantile(0.80)
print(f"\n  분위수 기준값: P20={p20:.4f}, P80={p80:.4f}")

def assign_quantile_group(cdi):
    if pd.isna(cdi):
        return "확인 필요"
    elif cdi < p20:
        return "수요 하위"
    elif cdi <= p80:
        return "수요 중위"
    else:
        return "수요 상위"

df["cdi_quantile_group"] = df["CDI"].apply(assign_quantile_group)

qgroup_order = ["수요 하위", "수요 중위", "수요 상위", "확인 필요"]
qgroup_counts = df["cdi_quantile_group"].value_counts()
print("  cdi_quantile_group 분포:")
for grp in qgroup_order:
    cnt = qgroup_counts.get(grp, 0)
    pct = cnt / n_rows_original * 100
    print(f"    {grp}: {cnt}개교 ({pct:.1f}%)")

# ════════════════════════════════════════════════════════════════════════════
# STEP 5. master_table.xlsx 저장 (기존 시트 보존, openpyxl 활용)
# ════════════════════════════════════════════════════════════════════════════
print("\n[STEP 5] master_table.xlsx 저장")

assert len(df) == n_rows_original, f"[ERROR] 행 수 변경됨: {n_rows_original} → {len(df)}"

# CDI 요약 / 결측 점검 시트용 데이터
summary_cols = [
    "school_code", "school_name", "sido", "sigungu",
    "demand_size_score", "counseling_use_score", "school_violence_risk_score",
    "CDI", "cdi_available_mean", "cdi_components_available",
    "cdi_missing_components", "cdi_level", "cdi_quantile_group",
]
cdi_summary_df = df[[c for c in summary_cols if c in df.columns]].copy()
cdi_missing_df = cdi_summary_df[
    (cdi_summary_df["cdi_components_available"] < 3) |
    cdi_summary_df["CDI"].isna()
].copy()

def write_sheet(ws, dataframe):
    """DataFrame을 openpyxl 시트에 헤더 포함 기록"""
    for c_idx, col_name in enumerate(dataframe.columns, start=1):
        ws.cell(row=1, column=c_idx, value=col_name)
    for r_idx, row in enumerate(dataframe.itertuples(index=False), start=2):
        for c_idx, val in enumerate(row, start=1):
            if isinstance(val, float) and np.isnan(val):
                ws.cell(row=r_idx, column=c_idx, value=None)
            else:
                ws.cell(row=r_idx, column=c_idx, value=val)

wb = load_workbook(MASTER_PATH)

# master_table 시트 교체
if "master_table" in wb.sheetnames:
    del wb["master_table"]
ws_master = wb.create_sheet("master_table", 0)
write_sheet(ws_master, df)

# cdi_summary 시트 추가/교체
for sname in ["cdi_summary", "cdi_missing_check"]:
    if sname in wb.sheetnames:
        del wb[sname]

ws_cdi_sum = wb.create_sheet("cdi_summary")
write_sheet(ws_cdi_sum, cdi_summary_df)

ws_cdi_mis = wb.create_sheet("cdi_missing_check")
if len(cdi_missing_df) > 0:
    write_sheet(ws_cdi_mis, cdi_missing_df)
else:
    ws_cdi_mis.cell(row=1, column=1, value="결측 없음")

wb.save(MASTER_PATH)
print(f"  저장 완료: {MASTER_PATH}")
print(f"  시트 목록: {wb.sheetnames}")

# ════════════════════════════════════════════════════════════════════════════
# STEP 6. CDI 전용 파일 생성
# ════════════════════════════════════════════════════════════════════════════
print("\n[STEP 6] CDI 전용 파일 생성")

CDI_TABLE_COLS = [
    "school_code", "school_name", "sido", "sigungu",
    "student_count",
    "demand_size_score",
    "avg_total_counseling_count_3yr",
    "counseling_count_per_student",
    "norm_avg_total_counseling_count_3yr",
    "norm_counseling_count_per_student",
    "counseling_use_score",
    "avg_victim_rate_3yr",
    "school_violence_risk_score",
    "CDI",
    "cdi_available_mean",
    "cdi_components_available",
    "cdi_missing_components",
    "cdi_level",
    "cdi_quantile_group",
]

# 실제로 존재하는 컬럼만 추출
cdi_table_df = df[[c for c in CDI_TABLE_COLS if c in df.columns]].copy()
missing_from_cdi_table = [c for c in CDI_TABLE_COLS if c not in df.columns]
if missing_from_cdi_table:
    print(f"  [WARN] CDI 전용 파일에서 누락된 컬럼: {missing_from_cdi_table}")

with pd.ExcelWriter(CDI_PATH, engine="openpyxl") as writer:
    cdi_table_df.to_excel(writer, sheet_name="CDI_table",        index=False)
    cdi_summary_df.to_excel(writer, sheet_name="cdi_summary",    index=False)
    cdi_missing_df.to_excel(writer, sheet_name="cdi_missing_check", index=False)

print(f"  저장 완료: {CDI_PATH}")

# ════════════════════════════════════════════════════════════════════════════
# STEP 7. CDI 요약 CSV 저장
# ════════════════════════════════════════════════════════════════════════════
print("\n[STEP 7] CDI 요약 CSV 저장")

summary_rows = []

def add_row(항목, 값):
    summary_rows.append({"항목": 항목, "값": 값})

add_row("전체 학교 수", n_rows_original)
add_row("CDI 평균",    round(df["CDI"].mean(), 4)   if df["CDI"].notna().any() else "N/A")
add_row("CDI 최솟값",  round(df["CDI"].min(), 4)    if df["CDI"].notna().any() else "N/A")
add_row("CDI 최댓값",  round(df["CDI"].max(), 4)    if df["CDI"].notna().any() else "N/A")
add_row("CDI 중앙값",  round(df["CDI"].median(), 4) if df["CDI"].notna().any() else "N/A")
add_row("CDI 결측 학교 수", df["CDI"].isna().sum())

for col in CDI_COLS:
    add_row(f"{col} 평균",    round(df[col].mean(), 4) if df[col].notna().any() else "N/A")
    add_row(f"{col} 결측 학교 수", df[col].isna().sum())

for lv in level_order:
    cnt = level_counts.get(lv, 0)
    pct = round(cnt / n_rows_original * 100, 1)
    add_row(f"cdi_level={lv} 학교 수", cnt)
    add_row(f"cdi_level={lv} 비율(%)", pct)

for grp in qgroup_order:
    cnt = qgroup_counts.get(grp, 0)
    pct = round(cnt / n_rows_original * 100, 1)
    add_row(f"cdi_quantile_group={grp} 학교 수", cnt)
    add_row(f"cdi_quantile_group={grp} 비율(%)", pct)

for n_comp in [0, 1, 2, 3]:
    cnt = (df["cdi_components_available"] == n_comp).sum()
    add_row(f"cdi_components_available={n_comp} 학교 수", cnt)

SUMMARY_CSV = TABLES_DIR / "cdi_summary_2025.csv"
pd.DataFrame(summary_rows).to_csv(SUMMARY_CSV, index=False, encoding="utf-8-sig")
print(f"  저장 완료: {SUMMARY_CSV}")

# ════════════════════════════════════════════════════════════════════════════
# STEP 8. 결측 점검 CSV 저장
# ════════════════════════════════════════════════════════════════════════════
print("\n[STEP 8] 결측 점검 CSV 저장")

MISSING_CSV = TABLES_DIR / "cdi_missing_check_2025.csv"
cdi_missing_df.to_csv(MISSING_CSV, index=False, encoding="utf-8-sig")
print(f"  저장 완료: {MISSING_CSV}")

# ════════════════════════════════════════════════════════════════════════════
# STEP 9. 방법론 문서 저장
# ════════════════════════════════════════════════════════════════════════════
print("\n[STEP 9] 방법론 문서 저장")

methodology_md = f"""# 상담수요지수(CDI) 산출 방법론

## 1. CDI의 목적
상담수요지수(CDI, Counseling Demand Index)는 경남 일반고 각 학교의 상담 수요 가능성을
정량화한 종합 지표이다. 학생 수 규모, 실제 상담 이용 수준, 학교폭력 피해 위험 수준을
종합하여 학교별 상담 수요 수준을 비교 가능한 단일 지표로 산출한다.

## 2. CDI 산출식
```
CDI = (demand_size_score + counseling_use_score + school_violence_risk_score) / 3
```
세 구성요소에 동일한 가중치(1/3)를 부여한다.

## 3. 구성요소별 산출 기준

### 3.1 demand_size_score (상담수요규모점수)
학교 전체 학생 수를 기준으로 구간 점수를 부여한다.

| 학생 수 | 점수 |
|---------|------|
| 500명 이상 | 1.0 |
| 250명 이상 500명 미만 | 0.6 |
| 250명 미만 | 0.3 |

### 3.2 counseling_use_score (실제상담이용점수)
2023~2025학년도 3개년 상담 건수를 바탕으로 다음 두 지표를 Min-Max 정규화한 후 평균한다.
- `norm_avg_total_counseling_count_3yr`: 3개년 평균 전체 상담 건수 정규화값
- `norm_counseling_count_per_student`: 학생 1인당 상담 건수 정규화값

```
counseling_use_score = (norm_avg_total_counseling_count_3yr + norm_counseling_count_per_student) / 2
```

### 3.3 school_violence_risk_score (학교폭력피해위험점수)
2023~2025학년도 학교폭력 실태조사 결과를 바탕으로 3개년 평균 피해율을 Min-Max 정규화한다.
```
school_violence_risk_score = (avg_victim_rate_3yr − min) / (max − min)
```

## 4. 최종 CDI 계산 기준
- 세 구성요소가 모두 존재할 때만 공식 CDI를 산출한다.
- 하나라도 결측이면 CDI = NaN으로 처리한다.

## 5. 결측 처리 기준
| 변수 | 기준 |
|------|------|
| CDI | 세 구성요소 모두 비결측일 때만 산출 |
| cdi_available_mean | 결측 제외 평균 (1~2개 구성요소 결측 시 참고용) |
| cdi_components_available | 비결측 구성요소 수 (0~3) |
| cdi_missing_components | 결측 구성요소명 목록 |

## 6. cdi_available_mean을 별도 생성한 이유
CDI가 결측인 학교에 대해 민감도 분석 및 내부 검토를 지원하기 위해 별도로 생성하였다.
`cdi_available_mean`은 공식 지표가 아니며, 결측 비율이 높은 학교에서는 해석에 주의가 필요하다.

## 7. cdi_level 그룹화 기준 (절대값 기준)
| CDI 구간 | 등급 |
|----------|------|
| 0.0 이상 0.3 미만 | 수요 낮음 |
| 0.3 이상 0.6 미만 | 수요 보통 |
| 0.6 이상 0.8 미만 | 수요 높음 |
| 0.8 이상 1.0 이하 | 수요 매우 높음 |
| CDI 결측 | 확인 필요 |

## 8. cdi_quantile_group 산출 기준 (분위수 기준)
CDI 결측값을 제외한 유효 학교 기준으로 분위수를 계산한다.
- 하위 20% (P20 미만): 수요 하위
- 중위 60% (P20 이상 P80 이하): 수요 중위
- 상위 20% (P80 초과): 수요 상위
- CDI 결측: 확인 필요

**주의**: 동일값이 P20 또는 P80 경계에 걸칠 경우 구간이 정확히 20/60/20으로
나뉘지 않을 수 있다. 이는 분위수 계산의 일반적인 특성으로 의도된 동작이다.

이번 산출에서의 분위수 기준값:
- P20 = {p20:.4f}
- P80 = {p80:.4f}

## 9. CDI 해석 방법
- CDI는 0과 1 사이의 값을 가지며, 값이 클수록 상담 수요 가능성이 높음을 의미한다.
- CDI가 높다고 해서 해당 학교에서 반드시 위기 학생이 많다는 것을 의미하지 않는다.
- CSI(상담공급지수)와 함께 해석할 때 의미 있는 시사점을 얻을 수 있다.
  - CDI 높음 & CSI 낮음: 수요 대비 공급 부족 → 우선 지원 검토 대상
  - CDI 낮음 & CSI 높음: 공급 충분, 수요 낮음 → 상대적으로 여유

## 10. 산출 결과 요약
- 대상 학교: {n_rows_original}개교
- CDI 비결측: {n_cdi_valid}개교
- CDI 평균: {df['CDI'].mean():.4f}
- CDI 범위: {df['CDI'].min():.4f} ~ {df['CDI'].max():.4f}

### cdi_level 분포
| 등급 | 학교 수 | 비율 |
|------|---------|------|
{"".join(f"| {lv} | {level_counts.get(lv, 0)}개교 | {level_counts.get(lv, 0)/n_rows_original*100:.1f}% |{chr(10)}" for lv in level_order)}

### cdi_quantile_group 분포
| 그룹 | 학교 수 | 비율 |
|------|---------|------|
{"".join(f"| {grp} | {qgroup_counts.get(grp, 0)}개교 | {qgroup_counts.get(grp, 0)/n_rows_original*100:.1f}% |{chr(10)}" for grp in qgroup_order)}

## 11. 한계
1. **학생 수 규모 지표의 한계**: demand_size_score는 학생 수를 기반으로 하므로
   상담 수요 가능성을 나타내는 규모 지표일 뿐 실제 위기 수준을 직접 의미하지 않는다.
2. **상담 건수 해석의 한계**: counseling_use_score의 기반인 상담 건수는 실제 수요뿐 아니라
   학교의 상담 접근성, 기록 방식, 운영 방식의 영향을 받을 수 있다.
3. **학교폭력 지표의 한계**: school_violence_risk_score는 피해 응답 자료 중심으로 구성되어
   가해·목격 응답은 반영하지 못한다.
4. **동일가중치 적용의 한계**: 세 구성요소에 동일한 가중치(1/3)를 적용하였으나,
   실제 상담 수요에 대한 각 요소의 기여도가 동일하다는 근거는 없다.
5. **직접 수요 지표가 아님**: CDI는 상담 수요 가능성을 나타내는 지표이며,
   실제 상담 필요 학생 수를 직접 의미하지 않는다.

## 12. 주의 사항
- 이 단계에서는 Priority Score를 산출하지 않는다.
- Priority Score는 추후 CDI - CSI 방식으로 별도 산출한다.
"""

METHODOLOGY_MD = DOCS_DIR / "cdi_methodology_2025.md"
METHODOLOGY_MD.write_text(methodology_md, encoding="utf-8")
print(f"  저장 완료: {METHODOLOGY_MD}")

# ════════════════════════════════════════════════════════════════════════════
# 최종 요약 출력
# ════════════════════════════════════════════════════════════════════════════
print("\n" + "=" * 60)
print("상담수요지수(CDI) 산출 완료")
print("=" * 60)
print(f"\n대상 학교: {n_rows_original}개교")
print(f"CDI 비결측: {n_cdi_valid}개교 / 결측: {n_cdi_missing}개교")
print(f"CDI 평균: {df['CDI'].mean():.4f}")
print(f"CDI 범위: {df['CDI'].min():.4f} ~ {df['CDI'].max():.4f}")
print(f"CDI 중앙값: {df['CDI'].median():.4f}")

print("\ncdi_level 분포:")
for lv in level_order:
    cnt = level_counts.get(lv, 0)
    print(f"  {lv}: {cnt}개교 ({cnt/n_rows_original*100:.1f}%)")

print("\ncdi_quantile_group 분포:")
for grp in qgroup_order:
    cnt = qgroup_counts.get(grp, 0)
    print(f"  {grp}: {cnt}개교 ({cnt/n_rows_original*100:.1f}%)")

print("\n생성 파일:")
print(f"  {MASTER_PATH}")
print(f"  {CDI_PATH}")
print(f"  {SUMMARY_CSV}")
print(f"  {MISSING_CSV}")
print(f"  {METHODOLOGY_MD}")
