"""
CDI 구성요소 중 demand_size_score(상담수요규모점수)와
counseling_use_score(실제상담이용점수)를 산출한다.

이번 단계에서는 school_violence_risk_score와 CDI 최종 점수는 산출하지 않는다.

입력 파일:
  data/processed/gyeongnam_high_schools_master_table.xlsx (master_table 시트)

출력 파일:
  data/processed/gyeongnam_high_schools_master_table.xlsx   (시트 추가 및 갱신)
  data/processed/gyeongnam_high_schools_CDI_partial.xlsx    (신규)
  outputs/tables/cdi_partial_summary_2025.csv
  outputs/tables/cdi_partial_missing_check_2025.csv
  docs/cdi_partial_methodology_2025.md
"""

import sys
import pathlib

import numpy as np
import pandas as pd
from openpyxl import load_workbook

# ── 경로 설정 ────────────────────────────────────────────────────────────────
ROOT          = pathlib.Path(__file__).resolve().parent.parent
PROCESSED_DIR = ROOT / "data" / "processed"
TABLES_DIR    = ROOT / "outputs" / "tables"
DOCS_DIR      = ROOT / "docs"

for d in [PROCESSED_DIR, TABLES_DIR, DOCS_DIR]:
    d.mkdir(parents=True, exist_ok=True)

# ════════════════════════════════════════════════════════════════════════════
# STEP 1. 입력 파일 로드
# ════════════════════════════════════════════════════════════════════════════
print("[STEP 1] 입력 파일 로드")

MASTER_PATH = PROCESSED_DIR / "gyeongnam_high_schools_master_table.xlsx"

if not MASTER_PATH.exists():
    print(f"[ERROR] 파일 없음: {MASTER_PATH}")
    sys.exit(1)

# master_table 시트 우선, 없으면 첫 번째 시트 사용
xl = pd.ExcelFile(MASTER_PATH)
if "master_table" in xl.sheet_names:
    sheet_used = "master_table"
else:
    sheet_used = xl.sheet_names[0]
    print(f"[INFO] 'master_table' 시트 없음 → '{sheet_used}' 시트 사용")

df = pd.read_excel(MASTER_PATH, sheet_name=sheet_used,
                   dtype={"school_code": str, "postcode": str})

print(f"  시트: '{sheet_used}'  |  {len(df)}행 × {len(df.columns)}열 로드 완료")

# 행 수 기준점 저장 (무결성 확인용)
original_row_count = len(df)

# ════════════════════════════════════════════════════════════════════════════
# STEP 2. 필수 변수 존재 여부 확인
# ════════════════════════════════════════════════════════════════════════════
print("\n[STEP 2] 필수 변수 확인")

def find_col(df, primary, candidates):
    """primary 변수명 우선 탐색, 없으면 후보 중 첫 번째 반환."""
    if primary in df.columns:
        return primary
    for c in candidates:
        if c in df.columns:
            print(f"  [INFO] '{primary}' 없음 → '{c}' 대체 사용")
            return c
    return None

id_map = {
    "school_code": find_col(df, "school_code", ["학교코드", "SCHUL_CODE"]),
    "school_name": find_col(df, "school_name", ["학교명", "SCHUL_NM"]),
    "sido"       : find_col(df, "sido",        ["시도"]),
    "sigungu"    : find_col(df, "sigungu",     ["시군구", "지역"]),
}

var_map = {
    "student_count"                    : find_col(df, "student_count",
        ["학생수", "total_student_count", "students"]),
    "avg_total_counseling_count_3yr"   : find_col(df, "avg_total_counseling_count_3yr",
        ["avg_total_counseling", "total_counseling", "counseling_count"]),
    "counseling_count_per_student"     : find_col(df, "counseling_count_per_student",
        ["counseling_per_student", "per_student_counseling"]),
    "norm_avg_total_counseling_count_3yr" : find_col(df, "norm_avg_total_counseling_count_3yr",
        ["norm_avg_total", "normalized_avg_counseling"]),
    "norm_counseling_count_per_student"   : find_col(df, "norm_counseling_count_per_student",
        ["norm_per_student", "normalized_per_student"]),
    "counseling_use_score"             : find_col(df, "counseling_use_score",
        ["use_score", "counseling_score"]),
}

# 필수 변수 결여 시 중단
missing_required = [k for k, v in var_map.items()
                    if v is None and k in ("student_count",
                                           "avg_total_counseling_count_3yr",
                                           "counseling_count_per_student")]
if missing_required:
    print(f"[ERROR] 다음 필수 변수를 찾지 못했습니다: {missing_required}")
    sys.exit(1)

all_vars = {**id_map, **var_map}
for name, found in all_vars.items():
    if found:
        miss = df[found].isna().sum()
        flag = "[OK]" if miss == 0 else "[WARN]"
        print(f"  {flag} {name:<45s} → '{found}'  결측={miss}")
    else:
        print(f"  [INFO] {name:<45s} → 없음 (이번 단계에서 산출 예정)")

# ════════════════════════════════════════════════════════════════════════════
# STEP 3. demand_size_score 생성
# student_count 기준 구간화: <250→1/3 / 250~499→2/3 / >=500→1.0 (균등 간격)
# ════════════════════════════════════════════════════════════════════════════
print("\n[STEP 3] demand_size_score 생성")

sc_col = var_map["student_count"]
sc     = pd.to_numeric(df[sc_col], errors="coerce")

# student_count == 0 확인 (실제 0과 결측 구분)
n_zero = int((sc == 0).sum())
if n_zero > 0:
    print(f"  [WARNING] student_count = 0인 학교 {n_zero}개 → 확인 필요로 기록")

def assign_demand_score(val) -> float:
    """student_count 구간화 점수 반환. 결측이면 NaN."""
    if pd.isna(val):
        return float("nan")
    v = float(val)
    if v < 250:
        return round(1 / 3, 4)   # 0.3333
    elif v < 500:
        return round(2 / 3, 4)   # 0.6667
    else:
        return 1.0

df["demand_size_score"] = sc.apply(assign_demand_score)

# 구간별 분포 출력
for score_val, label in [(0.3, "< 250"), (0.6, "250~499"), (1.0, ">= 500")]:
    cnt = int((df["demand_size_score"] == score_val).sum())
    print(f"  score={score_val} ({label:8s}): {cnt}개교  ({cnt/len(df)*100:.1f}%)")
print(f"  결측               : {int(df['demand_size_score'].isna().sum())}개교")
print(f"  평균               : {df['demand_size_score'].mean():.4f}")

# ════════════════════════════════════════════════════════════════════════════
# STEP 4. counseling_use_score 확인 및 메타 변수 생성
# 이미 script 03에서 동일 공식으로 산출된 값을 그대로 사용한다.
# norm_avg_total_counseling_count_3yr와 norm_counseling_count_per_student의
# 평균으로 counseling_use_score가 계산되어 있음을 검증한다.
# ════════════════════════════════════════════════════════════════════════════
print("\n[STEP 4] counseling_use_score 확인 및 메타 변수 생성")

norm1_col = var_map["norm_avg_total_counseling_count_3yr"]
norm2_col = var_map["norm_counseling_count_per_student"]
use_col   = var_map["counseling_use_score"]

comp_cols_use = [c for c in [norm1_col, norm2_col] if c is not None]

if use_col and norm1_col and norm2_col:
    # 기존 counseling_use_score와 재계산 값 비교 검증
    recalc = df[comp_cols_use].mean(axis=1, skipna=True)
    # 두 구성요소 모두 결측이면 NaN
    recalc[df[comp_cols_use].isna().all(axis=1)] = float("nan")

    max_diff = (df[use_col] - recalc).abs().max()
    if max_diff < 1e-9:
        print(f"  [OK] 기존 counseling_use_score 값 검증 완료 (최대 차이={max_diff:.2e})")
        print(f"       → 재계산 없이 기존 값 그대로 사용")
    else:
        print(f"  [INFO] 기존 값과 재계산 값 차이={max_diff:.6f} → 재계산 값으로 덮어씀")
        df[use_col] = recalc.round(6)

elif use_col is None:
    # counseling_use_score가 아예 없으면 직접 계산
    print("  [INFO] counseling_use_score 없음 → Min-Max 정규화 후 산출")

    def minmax(series: pd.Series) -> pd.Series:
        mn, mx = series.min(), series.max()
        if mx == mn:
            print(f"  [WARNING] min=max → 정규화 불가, NaN 처리")
            return pd.Series([float("nan")] * len(series), index=series.index)
        return (series - mn) / (mx - mn)

    raw1_col = var_map["avg_total_counseling_count_3yr"]
    raw2_col = var_map["counseling_count_per_student"]

    if "norm_avg_total_counseling_count_3yr" not in df.columns and raw1_col:
        df["norm_avg_total_counseling_count_3yr"] = minmax(
            pd.to_numeric(df[raw1_col], errors="coerce"))
        norm1_col = "norm_avg_total_counseling_count_3yr"

    if "norm_counseling_count_per_student" not in df.columns and raw2_col:
        df["norm_counseling_count_per_student"] = minmax(
            pd.to_numeric(df[raw2_col], errors="coerce"))
        norm2_col = "norm_counseling_count_per_student"

    comp_cols_use = [c for c in [norm1_col, norm2_col] if c is not None]
    df["counseling_use_score"] = df[comp_cols_use].mean(axis=1, skipna=True)
    df.loc[df[comp_cols_use].isna().all(axis=1), "counseling_use_score"] = float("nan")
    use_col = "counseling_use_score"

# 메타 변수 생성: 결측이 아닌 구성 지표 수와 결측 구성 지표명
df["counseling_use_score_components_available"] = (
    df[comp_cols_use].notna().sum(axis=1).astype(int)
)

def list_missing_use(row):
    missing = [c for c in comp_cols_use if pd.isna(row[c])]
    return "; ".join(missing) if missing else ""

df["counseling_use_score_missing_components"] = df.apply(list_missing_use, axis=1)

print(f"  counseling_use_score 평균  : {df[use_col].mean():.4f}")
print(f"  counseling_use_score 결측  : {df[use_col].isna().sum()}개교")
print(f"  components_available = 2   : {(df['counseling_use_score_components_available'] == 2).sum()}개교")
print(f"  components_available = 1   : {(df['counseling_use_score_components_available'] == 1).sum()}개교")
print(f"  components_available = 0   : {(df['counseling_use_score_components_available'] == 0).sum()}개교")

# ════════════════════════════════════════════════════════════════════════════
# STEP 5. 행 수 무결성 확인
# ════════════════════════════════════════════════════════════════════════════
assert len(df) == original_row_count, \
    f"[ERROR] 행 수 변경 감지: {original_row_count} → {len(df)}"
print(f"\n[STEP 5] 행 수 무결성 확인 완료: {len(df)}행 유지")

# ════════════════════════════════════════════════════════════════════════════
# STEP 6. 요약표 생성
# ════════════════════════════════════════════════════════════════════════════
print("[STEP 6] 요약표 생성")

def stat_row(label, series):
    s = pd.to_numeric(series, errors="coerce")
    return {"항목": label,
            "평균": round(s.mean(), 4) if s.notna().any() else "",
            "최솟값": round(s.min(), 4) if s.notna().any() else "",
            "최댓값": round(s.max(), 4) if s.notna().any() else "",
            "결측수": int(s.isna().sum())}

summary_rows = [
    {"항목": "전체 학교 수", "평균": len(df), "최솟값": "", "최댓값": "", "결측수": ""},
    {"항목": "--- demand_size_score ---",
     "평균": "", "최솟값": "", "최댓값": "", "결측수": ""},
    stat_row("demand_size_score", df["demand_size_score"]),
]

# demand_size_score 구간별 학교 수
for sv, lbl in [(0.3, "0.3 (student < 250)"),
                (0.6, "0.6 (250 <= student < 500)"),
                (1.0, "1.0 (student >= 500)")]:
    cnt = int((df["demand_size_score"] == sv).sum())
    summary_rows.append({"항목": f"  구간 {lbl}",
                         "평균": "", "최솟값": "",
                         "최댓값": f"{cnt}개교 ({cnt/len(df)*100:.1f}%)",
                         "결측수": ""})

summary_rows += [
    {"항목": "--- counseling_use_score ---",
     "평균": "", "최솟값": "", "최댓값": "", "결측수": ""},
    stat_row("avg_total_counseling_count_3yr",
             df[var_map["avg_total_counseling_count_3yr"]]),
    stat_row("counseling_count_per_student",
             df[var_map["counseling_count_per_student"]]),
    stat_row("norm_avg_total_counseling_count_3yr", df[norm1_col]),
    stat_row("norm_counseling_count_per_student",   df[norm2_col]),
    stat_row("counseling_use_score", df[use_col]),
]

for n_comp in [2, 1, 0]:
    cnt = int((df["counseling_use_score_components_available"] == n_comp).sum())
    summary_rows.append({
        "항목": f"  components_available = {n_comp}",
        "평균": "", "최솟값": "",
        "최댓값": f"{cnt}개교", "결측수": ""})

df_summary = pd.DataFrame(summary_rows)
summary_path = TABLES_DIR / "cdi_partial_summary_2025.csv"
df_summary.to_csv(summary_path, index=False, encoding="utf-8-sig")
print(f"  → 저장: {summary_path.name}")

# ════════════════════════════════════════════════════════════════════════════
# STEP 7. 결측 점검표 생성
# ════════════════════════════════════════════════════════════════════════════
print("[STEP 7] 결측 점검표 생성")

check_display = [
    "school_code", "school_name", "sido", "sigungu",
    var_map["student_count"],
    "demand_size_score",
    var_map["avg_total_counseling_count_3yr"],
    var_map["counseling_count_per_student"],
    norm1_col, norm2_col,
    use_col,
    "counseling_use_score_components_available",
    "counseling_use_score_missing_components",
]
check_display = list(dict.fromkeys(c for c in check_display if c))

sc_num = pd.to_numeric(df[var_map["student_count"]], errors="coerce")

check_masks = {
    "student_count 결측":                df[var_map["student_count"]].isna(),
    "student_count == 0":               sc_num == 0,
    "demand_size_score 결측":           df["demand_size_score"].isna(),
    "avg_total_counseling_count_3yr 결측":
        df[var_map["avg_total_counseling_count_3yr"]].isna(),
    "counseling_count_per_student 결측":
        df[var_map["counseling_count_per_student"]].isna(),
    "counseling_use_score 결측":        df[use_col].isna(),
    "components_available < 2":
        df["counseling_use_score_components_available"] < 2,
}

check_rows = []
for reason, mask in check_masks.items():
    subset = df[mask][check_display].copy()
    if len(subset) > 0:
        subset.insert(0, "check_reason", reason)
        check_rows.append(subset)

if check_rows:
    df_check = pd.concat(check_rows, ignore_index=True)
else:
    df_check = pd.DataFrame([{
        "check_reason": "이상 없음 — 전 146개교 정상 산출",
        **{c: "" for c in check_display}
    }])

check_path = TABLES_DIR / "cdi_partial_missing_check_2025.csv"
df_check.to_csv(check_path, index=False, encoding="utf-8-sig")
print(f"  → 저장: {check_path.name}  ({len(df_check)}건)")

# ════════════════════════════════════════════════════════════════════════════
# STEP 8. master_table.xlsx 갱신
# 기존 시트(master_table, csi_summary, csi_missing_check) 유지하면서
# master_table 시트 내용 갱신 + 새 시트 2개 추가
# ════════════════════════════════════════════════════════════════════════════
print("\n[STEP 8] master_table.xlsx 갱신")

# openpyxl로 기존 파일 로드 → 시트 교체·추가 후 저장
wb = load_workbook(MASTER_PATH)

def write_df_to_sheet(wb, sheet_name: str, df_data: pd.DataFrame):
    """DataFrame을 지정 시트에 쓴다. 시트가 있으면 교체, 없으면 신규 생성."""
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    # 헤더 쓰기
    for col_idx, col_name in enumerate(df_data.columns, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)
    # 데이터 쓰기
    for row_idx, row_data in enumerate(df_data.itertuples(index=False), start=2):
        for col_idx, val in enumerate(row_data, start=1):
            # NaN은 None으로 변환 (엑셀 빈 셀)
            ws.cell(row=row_idx, column=col_idx,
                    value=None if (isinstance(val, float) and val != val) else val)

write_df_to_sheet(wb, "master_table",          df)
write_df_to_sheet(wb, "cdi_partial_summary",   df_summary)
write_df_to_sheet(wb, "cdi_partial_missing_check", df_check)

# master_table을 첫 번째 시트로 이동
if wb.sheetnames[0] != "master_table":
    wb.move_sheet("master_table", offset=-wb.sheetnames.index("master_table"))

wb.save(MASTER_PATH)
print(f"  → 저장: {MASTER_PATH.name}")
print(f"    포함 시트: {wb.sheetnames}")
print(f"    master_table: {len(df)}행 × {len(df.columns)}열")

# ════════════════════════════════════════════════════════════════════════════
# STEP 9. CDI_partial 전용 엑셀 저장
# ════════════════════════════════════════════════════════════════════════════
print("\n[STEP 9] CDI_partial 전용 엑셀 저장")

partial_cols = [
    "school_code", "school_name", "sido", "sigungu",
    var_map["student_count"],
    "demand_size_score",
    var_map["avg_total_counseling_count_3yr"],
    var_map["counseling_count_per_student"],
    norm1_col, norm2_col,
    use_col,
    "counseling_use_score_components_available",
    "counseling_use_score_missing_components",
]
partial_cols = list(dict.fromkeys(c for c in partial_cols if c and c in df.columns))

df_partial = df[partial_cols].copy()

# 표준 변수명으로 열 이름 정규화
rename_map = {
    var_map["student_count"]                  : "student_count",
    var_map["avg_total_counseling_count_3yr"] : "avg_total_counseling_count_3yr",
    var_map["counseling_count_per_student"]   : "counseling_count_per_student",
    norm1_col                                 : "norm_avg_total_counseling_count_3yr",
    norm2_col                                 : "norm_counseling_count_per_student",
    use_col                                   : "counseling_use_score",
}
df_partial = df_partial.rename(columns={k: v for k, v in rename_map.items() if k != v})

partial_path = PROCESSED_DIR / "gyeongnam_high_schools_CDI_partial.xlsx"
with pd.ExcelWriter(partial_path, engine="openpyxl") as writer:
    df_partial.to_excel(writer, sheet_name="CDI_partial_table", index=False)
    df_summary.to_excel(writer, sheet_name="cdi_partial_summary",      index=False)
    df_check.to_excel(  writer, sheet_name="cdi_partial_missing_check", index=False)

print(f"  → 저장: {partial_path.name}")
print(f"    CDI_partial_table: {len(df_partial)}행 × {len(df_partial.columns)}열")
print(f"    포함 변수: {list(df_partial.columns)}")

# ════════════════════════════════════════════════════════════════════════════
# STEP 10. 방법론 문서 작성
# ════════════════════════════════════════════════════════════════════════════
print("\n[STEP 10] 방법론 문서 작성")

doc_lines = [
    "# CDI 구성요소 부분 산출 방법론 문서 (2025)",
    "",
    "## CDI 전체 구조",
    "",
    "```",
    "CDI = (demand_size_score + counseling_use_score + school_violence_risk_score) / 3",
    "```",
    "",
    "상담수요지수(CDI, Counseling Demand Index)는 학교의 상담 수요 수준을 세 가지",
    "구성요소로 측정하는 복합 지수이다.",
    "",
    "**이번 단계에서는 `demand_size_score`와 `counseling_use_score`만 산출하였다.**",
    "`school_violence_risk_score`는 별도 자료 수집 후 추후 산출 예정이며,",
    "CDI 최종 점수는 이번 단계에서 계산하지 않는다.",
    "",
    "## 1. demand_size_score (상담수요규모점수)",
    "",
    "### 목적",
    "",
    "학교의 재학생 수(student_count)를 기준으로 잠재적 상담 수요 규모를 점수화한다.",
    "규모가 클수록 절대적인 상담 수요가 많을 것으로 가정한다.",
    "",
    "### 구간화 기준",
    "",
    "| student_count | demand_size_score |",
    "|---------------|-------------------|",
    "| 250 미만 | 0.3 |",
    "| 250 이상 ~ 500 미만 | 0.6 |",
    "| 500 이상 | 1.0 |",
    "| 결측 | NaN |",
    "| 0 (실제 0값) | 확인 필요 |",
    "",
    "### 해석 시 주의",
    "",
    "- student_count는 상담 수요 가능성을 나타내는 규모 지표일 뿐,",
    "  실제 위기 수준이나 미충족 상담 수요를 직접 의미하지 않는다.",
    "- student_count는 CSI(상담공급지수) 산출 과정에서",
    "  students_per_counselor에 간접 반영된 바 있으나,",
    "  CDI에서는 수요 규모 지표로 별도 사용한다.",
    "",
    "## 2. counseling_use_score (실제상담이용점수)",
    "",
    "### 목적",
    "",
    "학생·학부모가 실제로 학교 상담을 이용한 수준을 점수화한다.",
    "상담 이용이 많을수록 상담 수요가 실제로 표출된 것으로 해석한다.",
    "",
    "### 사용 지표",
    "",
    "| 지표 | 변수명 | 의미 |",
    "|------|--------|------|",
    "| 3개년 평균 상담 건수 | `avg_total_counseling_count_3yr` | 2023~2025 학생·학부모 통합 상담 건수 3개년 평균 |",
    "| 학생 수 대비 상담 건수 | `counseling_count_per_student` | 재학생 1인당 상담 건수 (규모 보정 지표) |",
    "",
    "### Min-Max 정규화를 각각 적용한 이유",
    "",
    "두 지표는 단위가 다르므로(건수 vs 비율) 직접 평균할 수 없다.",
    "각 지표를 0~1 사이로 정규화한 후 평균하여 동등하게 반영한다.",
    "두 지표 모두 값이 높을수록 상담 이용 수준이 높은 정방향 지표이므로",
    "역정규화(1-x)하지 않는다.",
    "",
    "### 정규화 공식",
    "",
    "```",
    "normalized = (x - min) / (max - min)",
    "```",
    "",
    "- 결측값은 정규화 계산에서 제외한다.",
    "- max = min인 경우 해당 정규화 변수는 NaN으로 처리한다.",
    "",
    "### counseling_use_score 산출식",
    "",
    "```",
    "counseling_use_score = mean(",
    "    norm_avg_total_counseling_count_3yr,",
    "    norm_counseling_count_per_student",
    ")",
    "```",
    "",
    "### 결측 처리 기준",
    "",
    "| 상황 | counseling_use_score |",
    "|------|----------------------|",
    "| 두 지표 모두 존재 | 두 지표 평균 |",
    "| 한 지표만 존재 | 존재 지표만 사용 |",
    "| 둘 다 결측 | NaN |",
    "",
    "## 한계",
    "",
    "- **규모와 위기 수준의 혼동**: 학생 수가 많다고 해서 실제 상담 위기 수준이",
    "  반드시 높은 것은 아니다. demand_size_score는 잠재 규모 지표에 한정된다.",
    "- **상담 건수의 다중 원인**: 상담 건수는 실제 수요뿐 아니라 학교의 상담 접근성,",
    "  상담교사 역량, 기록 방식, 운영 체계의 영향을 받을 수 있다.",
    "- **미충족 수요 미반영**: `counseling_use_score`는 실제 이용된 상담 수준을",
    "  나타내며, 상담을 받지 못한 잠재적 수요(미충족 수요)는 직접 측정하지 못한다.",
    "- **school_violence_risk_score 미산출**: 학교폭력 위험 점수는",
    "  별도 공시 자료 수집 후 추후 산출 예정이며, CDI 최종 점수는 이후 단계에서 산출한다.",
]

doc_path = DOCS_DIR / "cdi_partial_methodology_2025.md"
doc_path.write_text("\n".join(doc_lines), encoding="utf-8")
print(f"  → 저장: {doc_path.name}")

# ════════════════════════════════════════════════════════════════════════════
# 최종 요약
# ════════════════════════════════════════════════════════════════════════════
print()
print("=" * 60)
print("최종 요약")
print("=" * 60)
print(f"  전체 학교 수                : {len(df)}개교")
print(f"  demand_size_score 산출 완료 : {df['demand_size_score'].notna().sum()}개교")
print(f"  demand_size_score 평균      : {df['demand_size_score'].mean():.4f}")
print(f"  counseling_use_score 평균   : {df[use_col].mean():.4f}")
print(f"  counseling_use_score 결측   : {df[use_col].isna().sum()}개교")
print(f"  신규 추가 변수              : demand_size_score,")
print(f"                               counseling_use_score_components_available,")
print(f"                               counseling_use_score_missing_components")
print("=" * 60)
print()
print("[생성/갱신 파일]")
print("  data/processed/gyeongnam_high_schools_master_table.xlsx   (갱신)")
print("  data/processed/gyeongnam_high_schools_CDI_partial.xlsx    (신규)")
print("  outputs/tables/cdi_partial_summary_2025.csv               (신규)")
print("  outputs/tables/cdi_partial_missing_check_2025.csv         (신규)")
print("  docs/cdi_partial_methodology_2025.md                      (신규)")
