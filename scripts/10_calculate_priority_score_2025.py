"""
우선지원점수(Priority Score) 산출, 우선지원등급 부여, 최종 요약 테이블 생성

Priority Score = CDI - CSI

이 단계에서는 기존 master table의 CSI와 CDI를 활용하여
우선지원점수와 등급을 산출하고 최종 요약 테이블을 생성한다.

입력 파일:
  data/processed/gyeongnam_high_schools_master_table.xlsx

출력 파일:
  data/processed/gyeongnam_high_schools_master_table.xlsx     (master_table 시트 갱신)
  data/processed/gyeongnam_high_schools_priority_summary.xlsx (신규)
  outputs/tables/priority_score_summary_2025.csv
  outputs/tables/priority_score_level_summary_2025.csv
  outputs/tables/priority_score_missing_check_2025.csv
  docs/priority_score_methodology_2025.md
"""

import sys
import pathlib

import numpy as np
import pandas as pd
from openpyxl import load_workbook

# ── 경로 설정 ─────────────────────────────────────────────────────────────────
ROOT          = pathlib.Path(__file__).resolve().parent.parent
PROCESSED_DIR = ROOT / "data" / "processed"
TABLES_DIR    = ROOT / "outputs" / "tables"
DOCS_DIR      = ROOT / "docs"

for d in [PROCESSED_DIR, TABLES_DIR, DOCS_DIR]:
    d.mkdir(parents=True, exist_ok=True)

MASTER_PATH   = PROCESSED_DIR / "gyeongnam_high_schools_master_table.xlsx"
PRIORITY_PATH = PROCESSED_DIR / "gyeongnam_high_schools_priority_summary.xlsx"

# ════════════════════════════════════════════════════════════════════════════
# STEP 1. 입력 파일 로드
# ════════════════════════════════════════════════════════════════════════════
print("[STEP 1] 입력 파일 로드")

if not MASTER_PATH.exists():
    print(f"[ERROR] 파일 없음: {MASTER_PATH}")
    sys.exit(1)

xl = pd.ExcelFile(MASTER_PATH)
if "master_table" in xl.sheet_names:
    sheet_used = "master_table"
else:
    sheet_used = xl.sheet_names[0]
    print(f"  [WARN] 'master_table' 시트 없음 — '{sheet_used}' 시트 사용")

master_df = pd.read_excel(MASTER_PATH, sheet_name=sheet_used)
n_rows_original = len(master_df)
print(f"  로드 완료: {n_rows_original}행 × {len(master_df.columns)}열 (시트: {sheet_used})")

# ════════════════════════════════════════════════════════════════════════════
# STEP 2. 필수 변수 확인
# ════════════════════════════════════════════════════════════════════════════
print("\n[STEP 2] 필수 변수 확인")

ID_COLS  = ["school_code", "school_name", "sido", "sigungu"]
CSI_SUB  = ["counseling_staff_supply_score", "wee_class_score", "wee_center_access_score"]
CDI_SUB  = ["demand_size_score", "counseling_use_score", "school_violence_risk_score"]
IDX_COLS = ["CSI", "CDI"]
ALL_REQUIRED = ID_COLS + CSI_SUB + CDI_SUB + IDX_COLS

missing_vars = [c for c in ALL_REQUIRED if c not in master_df.columns]
if missing_vars:
    print(f"  [ERROR] 누락 변수: {missing_vars}")
    print("  유사 변수명 탐색:")
    for mv in missing_vars:
        candidates = [c for c in master_df.columns if mv[:5].lower() in c.lower()]
        print(f"    '{mv}' 후보: {candidates if candidates else '없음'}")
    sys.exit(1)

print("  필수 변수 전체 확인 완료")
for col in IDX_COLS + CSI_SUB + CDI_SUB:
    n_valid = master_df[col].notna().sum()
    mean_val = master_df[col].mean()
    print(f"    {col}: {n_valid}/{n_rows_original} 비결측, 평균={mean_val:.4f}")

# ════════════════════════════════════════════════════════════════════════════
# STEP 3. 우선지원점수 산출
# ════════════════════════════════════════════════════════════════════════════
print("\n[STEP 3] 우선지원점수(Priority Score) 산출")

df = master_df.copy()

# 기존 컬럼 중복 제거
for col in ["priority_score", "priority_missing_reason", "priority_level",
            "priority_level_fixed"]:
    if col in df.columns:
        df.drop(columns=[col], inplace=True)

cdi_ok = df["CDI"].notna()
csi_ok = df["CSI"].notna()

# Priority Score = CDI - CSI (둘 다 존재할 때만)
df["priority_score"] = np.where(
    cdi_ok & csi_ok,
    df["CDI"] - df["CSI"],
    np.nan,
)

# 결측 사유 기록
def missing_reason(row):
    cdi_miss = pd.isna(row["CDI"])
    csi_miss = pd.isna(row["CSI"])
    if cdi_miss and csi_miss:
        return "CDI 및 CSI 결측"
    elif cdi_miss:
        return "CDI 결측"
    elif csi_miss:
        return "CSI 결측"
    else:
        return "없음"

df["priority_missing_reason"] = df[["CDI", "CSI"]].apply(missing_reason, axis=1)

n_valid   = df["priority_score"].notna().sum()
n_missing = df["priority_score"].isna().sum()
ps        = df["priority_score"].dropna()

print(f"  priority_score 산출: {n_valid}개교 / 결측: {n_missing}개교")
print(f"  priority_score 범위: {ps.min():.4f} ~ {ps.max():.4f}")
print(f"  priority_score 평균: {ps.mean():.4f}  중앙값: {ps.median():.4f}")
print(f"  0 초과(수요>공급): {(ps > 0).sum()}개교  /  0 이하(공급≥수요): {(ps <= 0).sum()}개교")

# ════════════════════════════════════════════════════════════════════════════
# STEP 4. 우선지원등급 부여 (분위수 기반 공식 등급)
# ════════════════════════════════════════════════════════════════════════════
print("\n[STEP 4] 우선지원등급 부여 (분위수 기반)")

p20 = ps.quantile(0.20)
p80 = ps.quantile(0.80)
p90 = ps.quantile(0.90)

print(f"  분위수 기준값: P20={p20:.4f}, P80={p80:.4f}, P90={p90:.4f}")

def assign_priority_level(score):
    if pd.isna(score):
        return "확인 필요"
    elif score > p90:
        return "최우선 지원"
    elif score > p80:
        return "우선 지원"
    elif score > p20:
        return "모니터링"
    else:
        return "안정"

df["priority_level"] = df["priority_score"].apply(assign_priority_level)

level_order = ["최우선 지원", "우선 지원", "모니터링", "안정", "확인 필요"]
level_counts = df["priority_level"].value_counts()
print("  priority_level 분포:")
for lv in level_order:
    cnt = level_counts.get(lv, 0)
    print(f"    {lv}: {cnt}개교 ({cnt/n_rows_original*100:.1f}%)")

# ════════════════════════════════════════════════════════════════════════════
# STEP 5. 고정 기준 보조 등급 생성
# ════════════════════════════════════════════════════════════════════════════
print("\n[STEP 5] 고정 기준 보조 등급 생성")

def assign_priority_level_fixed(score):
    if pd.isna(score):
        return "확인 필요"
    elif score >= 0.30:
        return "최우선 지원 검토"
    elif score >= 0.15:
        return "우선 지원 검토"
    elif score >= 0.00:
        return "모니터링"
    else:
        return "안정"

df["priority_level_fixed"] = df["priority_score"].apply(assign_priority_level_fixed)

fixed_order = ["최우선 지원 검토", "우선 지원 검토", "모니터링", "안정", "확인 필요"]
fixed_counts = df["priority_level_fixed"].value_counts()
print("  priority_level_fixed 분포:")
for lv in fixed_order:
    cnt = fixed_counts.get(lv, 0)
    print(f"    {lv}: {cnt}개교 ({cnt/n_rows_original*100:.1f}%)")

# ════════════════════════════════════════════════════════════════════════════
# STEP 6. master_table.xlsx 저장 (기존 시트 보존)
# ════════════════════════════════════════════════════════════════════════════
print("\n[STEP 6] master_table.xlsx 저장")

assert len(df) == n_rows_original, f"[ERROR] 행 수 변경됨: {n_rows_original} → {len(df)}"

# 요약 / 결측 점검 시트용 데이터
SUMMARY_COLS = [
    "school_code", "school_name", "sido", "sigungu",
    "CSI", "CDI", "priority_score",
    "priority_level", "priority_level_fixed", "priority_missing_reason",
]
ps_summary_df = df[[c for c in SUMMARY_COLS if c in df.columns]].copy()

MISSING_COLS = [
    "school_code", "school_name", "sido", "sigungu",
    "counseling_staff_supply_score", "wee_class_score", "wee_center_access_score",
    "demand_size_score", "counseling_use_score", "school_violence_risk_score",
    "CSI", "CDI", "priority_score", "priority_missing_reason",
]
ps_missing_df = df[
    df["priority_missing_reason"] != "없음"
][[c for c in MISSING_COLS if c in df.columns]].copy()

def write_sheet(ws, dataframe):
    """DataFrame을 openpyxl 워크시트에 헤더 포함 기록"""
    for c_idx, col_name in enumerate(dataframe.columns, start=1):
        ws.cell(row=1, column=c_idx, value=col_name)
    for r_idx, row in enumerate(dataframe.itertuples(index=False), start=2):
        for c_idx, val in enumerate(row, start=1):
            if isinstance(val, float) and np.isnan(val):
                ws.cell(row=r_idx, column=c_idx, value=None)
            else:
                ws.cell(row=r_idx, column=c_idx, value=val)

wb = load_workbook(MASTER_PATH)

# master_table 시트 교체
if "master_table" in wb.sheetnames:
    del wb["master_table"]
ws_master = wb.create_sheet("master_table", 0)
write_sheet(ws_master, df)

# priority 관련 시트 추가/교체
for sname in ["priority_score_summary", "priority_score_missing_check"]:
    if sname in wb.sheetnames:
        del wb[sname]

ws_ps_sum = wb.create_sheet("priority_score_summary")
write_sheet(ws_ps_sum, ps_summary_df)

ws_ps_mis = wb.create_sheet("priority_score_missing_check")
if len(ps_missing_df) > 0:
    write_sheet(ws_ps_mis, ps_missing_df)
else:
    ws_ps_mis.cell(row=1, column=1, value="결측 없음")

wb.save(MASTER_PATH)
print(f"  저장 완료: {MASTER_PATH}")
print(f"  시트 목록: {wb.sheetnames}")

# ════════════════════════════════════════════════════════════════════════════
# STEP 7. 최종 요약용 테이블 생성 (priority_summary.xlsx)
# ════════════════════════════════════════════════════════════════════════════
print("\n[STEP 7] 최종 요약 테이블 생성")

FINAL_COLS = [
    "school_code", "school_name", "sido", "sigungu",
    # CSI 하위
    "counseling_staff_supply_score", "wee_class_score", "wee_center_access_score",
    # CDI 하위
    "demand_size_score", "counseling_use_score", "school_violence_risk_score",
    # 최종 지수
    "CSI", "CDI",
    "priority_score", "priority_level", "priority_level_fixed",
    "priority_missing_reason",
]
final_df = df[[c for c in FINAL_COLS if c in df.columns]].copy()
missing_from_final = [c for c in FINAL_COLS if c not in df.columns]
if missing_from_final:
    print(f"  [WARN] 요약 테이블 누락 컬럼: {missing_from_final}")

# priority_score_level_summary 시트용 — 등급별 분포
level_summary_rows = []
for lv in level_order:
    cnt = level_counts.get(lv, 0)
    level_summary_rows.append({
        "구분": "priority_level (분위수)",
        "등급": lv,
        "학교 수": cnt,
        "비율(%)": round(cnt / n_rows_original * 100, 1),
    })
for lv in fixed_order:
    cnt = fixed_counts.get(lv, 0)
    level_summary_rows.append({
        "구분": "priority_level_fixed (고정기준)",
        "등급": lv,
        "학교 수": cnt,
        "비율(%)": round(cnt / n_rows_original * 100, 1),
    })

# 시군구별 priority_score 평균 및 최우선·우선 지원 학교 수
sigungu_grp = df.groupby("sigungu").agg(
    학교수        = ("school_code", "count"),
    priority_score_평균 = ("priority_score", "mean"),
    최우선지원_수  = ("priority_level", lambda x: (x == "최우선 지원").sum()),
    우선지원_수    = ("priority_level", lambda x: (x == "우선 지원").sum()),
).reset_index().round(4)

level_summary_df = pd.DataFrame(level_summary_rows)

with pd.ExcelWriter(PRIORITY_PATH, engine="openpyxl") as writer:
    final_df.to_excel(writer,         sheet_name="priority_summary_table",     index=False)
    ps_summary_df.to_excel(writer,    sheet_name="priority_score_summary",      index=False)
    level_summary_df.to_excel(writer, sheet_name="priority_score_level_summary", index=False)
    if len(ps_missing_df) > 0:
        ps_missing_df.to_excel(writer, sheet_name="priority_score_missing_check", index=False)
    else:
        pd.DataFrame({"결과": ["결측 없음"]}).to_excel(
            writer, sheet_name="priority_score_missing_check", index=False)

print(f"  저장 완료: {PRIORITY_PATH}")

# ════════════════════════════════════════════════════════════════════════════
# STEP 8. 우선지원점수 요약 CSV 저장
# ════════════════════════════════════════════════════════════════════════════
print("\n[STEP 8] 우선지원점수 요약 CSV 저장")

summary_rows = []
def add(항목, 값):
    summary_rows.append({"항목": 항목, "값": 값})

add("전체 학교 수",           n_rows_original)
add("priority_score 산출 학교 수", n_valid)
add("priority_score 결측 학교 수", n_missing)
add("priority_score 평균",    round(ps.mean(), 4))
add("priority_score 최솟값",  round(ps.min(),  4))
add("priority_score 최댓값",  round(ps.max(),  4))
add("priority_score 중앙값",  round(ps.median(), 4))
add("CSI 평균",               round(df["CSI"].mean(), 4))
add("CDI 평균",               round(df["CDI"].mean(), 4))
add("P20 기준값",             round(p20, 4))
add("P80 기준값",             round(p80, 4))
add("P90 기준값",             round(p90, 4))

SUMMARY_CSV = TABLES_DIR / "priority_score_summary_2025.csv"
base_df = pd.DataFrame(summary_rows)

# 상위 10개교
top10 = (
    df[df["priority_score"].notna()]
    .nlargest(10, "priority_score")
    [["school_name", "sigungu", "CSI", "CDI", "priority_score",
      "priority_level", "priority_level_fixed"]]
    .reset_index(drop=True)
)
top10.index += 1
top10.index.name = "순위"

# 하위 10개교
bot10 = (
    df[df["priority_score"].notna()]
    .nsmallest(10, "priority_score")
    [["school_name", "sigungu", "CSI", "CDI", "priority_score",
      "priority_level", "priority_level_fixed"]]
    .reset_index(drop=True)
)
bot10.index += 1
bot10.index.name = "순위"

# 하나의 CSV에 순서대로 저장
with open(SUMMARY_CSV, "w", encoding="utf-8-sig", newline="") as f:
    base_df.to_csv(f, index=False)
    f.write("\n[priority_score 상위 10개교]\n")
    top10.to_csv(f)
    f.write("\n[priority_score 하위 10개교]\n")
    bot10.to_csv(f)

print(f"  저장 완료: {SUMMARY_CSV}")

# ════════════════════════════════════════════════════════════════════════════
# STEP 9. 우선지원등급 요약 CSV 저장
# ════════════════════════════════════════════════════════════════════════════
print("\n[STEP 9] 우선지원등급 요약 CSV 저장")

LEVEL_CSV = TABLES_DIR / "priority_score_level_summary_2025.csv"
with open(LEVEL_CSV, "w", encoding="utf-8-sig", newline="") as f:
    level_summary_df.to_csv(f, index=False)
    f.write("\n[시군구별 priority_score 평균 및 최우선·우선 지원 학교 수]\n")
    sigungu_grp.to_csv(f, index=False)

print(f"  저장 완료: {LEVEL_CSV}")

# ════════════════════════════════════════════════════════════════════════════
# STEP 10. 결측 점검 CSV 저장
# ════════════════════════════════════════════════════════════════════════════
print("\n[STEP 10] 결측 점검 CSV 저장")

MISSING_CSV = TABLES_DIR / "priority_score_missing_check_2025.csv"
if len(ps_missing_df) > 0:
    ps_missing_df.to_csv(MISSING_CSV, index=False, encoding="utf-8-sig")
else:
    pd.DataFrame({"결과": ["결측 없음"]}).to_csv(MISSING_CSV, index=False, encoding="utf-8-sig")

print(f"  저장 완료: {MISSING_CSV}")

# ════════════════════════════════════════════════════════════════════════════
# STEP 11. 방법론 문서 저장
# ════════════════════════════════════════════════════════════════════════════
print("\n[STEP 11] 방법론 문서 저장")

methodology_md = f"""# 우선지원점수(Priority Score) 산출 방법론

## 1. Priority Score의 목적
우선지원점수(Priority Score)는 경남 일반고 각 학교의 상담 수요 대비 공급 충족 수준을
정량화하여 정책적 지원 우선순위를 도출하기 위한 지표이다.
상담수요지수(CDI)에서 상담공급지수(CSI)를 뺀 값으로 산출하며,
값이 클수록 수요 대비 공급이 부족한 학교로 해석한다.

## 2. Priority Score 산출식
```
Priority Score = CDI - CSI
```

## 3. CSI 구성요소
| 변수 | 설명 |
|------|------|
| `counseling_staff_supply_score` | 전문상담인력 공급 점수 (학생 1인당 상담 인력 기반 구간 점수) |
| `wee_class_score` | Wee 클래스 운영 여부 (운영=1.0, 미운영=0.0) |
| `wee_center_access_score` | Wee 센터 접근성 점수 (최근접 Wee 센터까지의 직선거리 기반 구간 점수) |

```
CSI = (counseling_staff_supply_score + wee_class_score + wee_center_access_score) / 3
```

## 4. CDI 구성요소
| 변수 | 설명 |
|------|------|
| `demand_size_score` | 상담수요규모점수 (학생 수 기반 구간 점수) |
| `counseling_use_score` | 실제상담이용점수 (3개년 상담 건수 Min-Max 정규화 평균) |
| `school_violence_risk_score` | 학교폭력 피해 위험 점수 (3개년 평균 피해율 Min-Max 정규화) |

```
CDI = (demand_size_score + counseling_use_score + school_violence_risk_score) / 3
```

## 5. Priority Score 해석 방식
| 값 범위 | 해석 |
|---------|------|
| 양수 (> 0) | 수요가 공급보다 상대적으로 높은 상태 → 지원 우선 검토 필요 |
| 0 근접 | 수요와 공급이 비슷한 균형 상태 |
| 음수 (< 0) | 공급이 수요보다 상대적으로 충분한 상태 |

## 6. priority_level 분위수 기반 등급화 기준
결측값을 제외한 유효 학교 기준으로 분위수를 계산한다.
학교 간 **상대 비교** 기반의 공식 등급이다.

| 구간 | 등급 |
|------|------|
| P90 초과 (상위 10%) | 최우선 지원 |
| P80 초과 ~ P90 이하 (상위 10~20%) | 우선 지원 |
| P20 초과 ~ P80 이하 (중위 60%) | 모니터링 |
| P20 이하 (하위 20%) | 안정 |
| 결측 | 확인 필요 |

이번 산출의 분위수 기준값:
- P20 = {p20:.4f}
- P80 = {p80:.4f}
- P90 = {p90:.4f}

**주의**: 동일값이 경계에 걸릴 경우 정확히 10/10/60/20%로 나뉘지 않을 수 있다.
이는 분위수 계산의 일반적인 특성으로 의도된 동작이다.

## 7. priority_level_fixed 고정 기준 보조 등급
해석 편의를 위해 절대값 기준의 보조 등급을 병행 제공한다.
`priority_level`이 공식 등급이며, `priority_level_fixed`는 참고용이다.

| priority_score 구간 | 등급 |
|--------------------|------|
| ≥ 0.30 | 최우선 지원 검토 |
| 0.15 이상 0.30 미만 | 우선 지원 검토 |
| 0.00 이상 0.15 미만 | 모니터링 |
| 0.00 미만 | 안정 |
| 결측 | 확인 필요 |

## 8. 우선지원등급의 의미
우선지원등급은 정책적 지원 검토의 우선순위를 나타내는 지표이며,
실제 지원 확정을 의미하지 않는다.
현장 방문, 담당 교사 의견, 개별 학생 사례 등을 종합하여
최종 지원 여부를 결정해야 한다.

## 9. 결측 처리 기준
| 조건 | priority_score | priority_missing_reason |
|------|----------------|-------------------------|
| CDI, CSI 모두 존재 | CDI - CSI | 없음 |
| CDI만 결측 | NaN | CDI 결측 |
| CSI만 결측 | NaN | CSI 결측 |
| CDI, CSI 모두 결측 | NaN | CDI 및 CSI 결측 |

## 10. 산출 결과 요약
- 대상 학교: {n_rows_original}개교
- priority_score 산출: {n_valid}개교 / 결측: {n_missing}개교
- priority_score 평균: {ps.mean():.4f}
- priority_score 범위: {ps.min():.4f} ~ {ps.max():.4f}
- priority_score 중앙값: {ps.median():.4f}
- 0 초과(수요>공급): {(ps > 0).sum()}개교
- 0 이하(공급≥수요): {(ps <= 0).sum()}개교

### priority_level 분포 (분위수 기반)
| 등급 | 학교 수 | 비율 |
|------|---------|------|
{"".join(f"| {lv} | {level_counts.get(lv,0)}개교 | {level_counts.get(lv,0)/n_rows_original*100:.1f}% |{chr(10)}" for lv in level_order)}
### priority_level_fixed 분포 (고정 기준)
| 등급 | 학교 수 | 비율 |
|------|---------|------|
{"".join(f"| {lv} | {fixed_counts.get(lv,0)}개교 | {fixed_counts.get(lv,0)/n_rows_original*100:.1f}% |{chr(10)}" for lv in fixed_order)}
## 11. 한계
1. **동일가중치 기반의 한계**: CSI와 CDI 모두 구성요소에 동일가중치(1/3)를 적용하였으며,
   각 요소의 실제 기여도가 동일하다는 근거는 없다.
2. **행정적 우선검토 지표**: Priority Score는 상담 필요성을 직접 확정하는 지표가 아니라
   정책 검토의 우선순위를 나타내는 참고 지표이다.
3. **직접 반영 불가 요소**: 학교별 상담의 질, 지역 교통 여건, 개별 위기 사례 등은
   현재 지표 구조에서 직접 반영하지 못한다.
4. **데이터 품질 의존**: 데이터 공개 범위와 수작업 입력 자료의 정확성에 영향을 받을 수 있다.
   특히 상담 건수 데이터는 학교별 기록 방식에 따라 편차가 있을 수 있다.
"""

METHODOLOGY_MD = DOCS_DIR / "priority_score_methodology_2025.md"
METHODOLOGY_MD.write_text(methodology_md, encoding="utf-8")
print(f"  저장 완료: {METHODOLOGY_MD}")

# ════════════════════════════════════════════════════════════════════════════
# 최종 요약 출력
# ════════════════════════════════════════════════════════════════════════════
print("\n" + "=" * 60)
print("우선지원점수(Priority Score) 산출 완료")
print("=" * 60)
print(f"\n대상 학교: {n_rows_original}개교")
print(f"priority_score 산출: {n_valid}개교 / 결측: {n_missing}개교")
print(f"priority_score 평균: {ps.mean():.4f}")
print(f"priority_score 범위: {ps.min():.4f} ~ {ps.max():.4f}")

print("\npriority_level 분포 (분위수 기반):")
for lv in level_order:
    cnt = level_counts.get(lv, 0)
    print(f"  {lv}: {cnt}개교 ({cnt/n_rows_original*100:.1f}%)")

print("\npriority_level_fixed 분포 (고정 기준):")
for lv in fixed_order:
    cnt = fixed_counts.get(lv, 0)
    print(f"  {lv}: {cnt}개교 ({cnt/n_rows_original*100:.1f}%)")

print(f"\npriority_score 상위 10개교:")
print(top10[["school_name", "sigungu", "priority_score", "priority_level"]].to_string())

print("\n생성 파일:")
for p in [MASTER_PATH, PRIORITY_PATH, SUMMARY_CSV, LEVEL_CSV, MISSING_CSV, METHODOLOGY_MD]:
    print(f"  {p}")
