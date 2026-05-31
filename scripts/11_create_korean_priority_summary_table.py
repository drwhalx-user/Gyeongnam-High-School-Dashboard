"""
최종 우선지원 요약 테이블의 표시용 한글 변수명 파일 생성

입력 파일:
  data/processed/gyeongnam_high_schools_priority_summary.xlsx (priority_summary_table 시트)

출력 파일:
  data/processed/gyeongnam_high_schools_priority_summary_korean.xlsx
    - 최종요약표 시트: 한글 표시명 + 서식 적용
    - 변수설명 시트: 변수 메타데이터
"""

import pathlib

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── 경로 설정 ─────────────────────────────────────────────────────────────────
ROOT          = pathlib.Path(__file__).resolve().parent.parent
PROCESSED_DIR = ROOT / "data" / "processed"
PROCESSED_DIR.mkdir(parents=True, exist_ok=True)

INPUT_PATH  = PROCESSED_DIR / "gyeongnam_high_schools_priority_summary.xlsx"
OUTPUT_PATH = PROCESSED_DIR / "gyeongnam_high_schools_priority_summary_korean.xlsx"

# ════════════════════════════════════════════════════════════════════════════
# STEP 1. 입력 파일 로드
# ════════════════════════════════════════════════════════════════════════════
print("[STEP 1] 입력 파일 로드")

if not INPUT_PATH.exists():
    raise FileNotFoundError(f"파일 없음: {INPUT_PATH}")

xl = pd.ExcelFile(INPUT_PATH)
if "priority_summary_table" in xl.sheet_names:
    sheet_used = "priority_summary_table"
else:
    sheet_used = xl.sheet_names[0]
    print(f"  [WARN] 'priority_summary_table' 없음 - '{sheet_used}' 시트 사용")

df = pd.read_excel(INPUT_PATH, sheet_name=sheet_used, dtype={"school_code": str})
n_rows_original = len(df)
print(f"  로드 완료: {n_rows_original}행 × {len(df.columns)}열 (시트: {sheet_used})")
print(f"  컬럼: {list(df.columns)}")

# ════════════════════════════════════════════════════════════════════════════
# STEP 2. 변수명 매핑 정의
# ════════════════════════════════════════════════════════════════════════════
print("\n[STEP 2] 변수명 매핑 정의")

# 원래 변수명 → 한글 표시명
COL_MAP = {
    "school_code":                  "학교코드",
    "school_name":                  "학교명",
    "sido":                         "시도",
    "sigungu":                      "시군구",
    "counseling_staff_supply_score":"상담인력 공급 점수",
    "wee_class_score":              "Wee클래스 운영 점수",
    "wee_center_access_score":      "Wee센터 접근성 점수",
    "CSI":                          "상담공급지수",
    "demand_size_score":            "상담수요 규모 점수",
    "counseling_use_score":         "실제 상담 이용 점수",
    "school_violence_risk_score":   "학교폭력 위험 점수",
    "CDI":                          "상담수요지수",
    "priority_score":               "우선지원점수",
    "priority_level":               "우선지원등급",
    "priority_level_fixed":         "고정기준 우선지원등급",
    "priority_missing_reason":      "우선지원점수 결측 사유",
}

# 출력 열 순서 (한글명 기준)
COL_ORDER_KR = [
    "학교코드", "학교명", "시도", "시군구",
    "상담인력 공급 점수", "Wee클래스 운영 점수", "Wee센터 접근성 점수", "상담공급지수",
    "상담수요 규모 점수", "실제 상담 이용 점수", "학교폭력 위험 점수", "상담수요지수",
    "우선지원점수", "우선지원등급", "고정기준 우선지원등급", "우선지원점수 결측 사유",
]

# 누락 변수 확인
missing_vars = [eng for eng, kr in COL_MAP.items() if eng not in df.columns]
present_map  = {eng: kr for eng, kr in COL_MAP.items() if eng in df.columns}

if missing_vars:
    print(f"  [INFO] 입력 파일에 없는 변수 {len(missing_vars)}개 (변환 제외):")
    for v in missing_vars:
        print(f"    - {v} → {COL_MAP[v]}")
else:
    print("  모든 매핑 변수 존재 확인 완료")

# ════════════════════════════════════════════════════════════════════════════
# STEP 3. 컬럼명 변경 및 열 순서 정리
# ════════════════════════════════════════════════════════════════════════════
print("\n[STEP 3] 컬럼명 변경 및 열 순서 정리")

df_kr = df.rename(columns=present_map)

# 실제로 존재하는 열만 순서대로 추출
final_cols = [kr for kr in COL_ORDER_KR if kr in df_kr.columns]
df_kr = df_kr[final_cols].copy()
print(f"  변환 후 컬럼 수: {len(df_kr.columns)}열")
print(f"  열 순서: {list(df_kr.columns)}")

# ════════════════════════════════════════════════════════════════════════════
# STEP 4. 점수형 변수 반올림 및 학교코드 문자열 처리
# ════════════════════════════════════════════════════════════════════════════
print("\n[STEP 4] 점수 반올림 및 형식 정리")

SCORE_COLS_KR = [
    "상담인력 공급 점수", "Wee클래스 운영 점수", "Wee센터 접근성 점수", "상담공급지수",
    "상담수요 규모 점수", "실제 상담 이용 점수", "학교폭력 위험 점수", "상담수요지수",
    "우선지원점수",
]
for col in SCORE_COLS_KR:
    if col in df_kr.columns:
        df_kr[col] = df_kr[col].round(3)

# 학교코드 문자열 유지
if "학교코드" in df_kr.columns:
    df_kr["학교코드"] = df_kr["학교코드"].astype(str).str.strip()

print(f"  점수형 열 반올림 완료: {[c for c in SCORE_COLS_KR if c in df_kr.columns]}")

# ════════════════════════════════════════════════════════════════════════════
# STEP 5. 정렬: 우선지원등급 우선 → 우선지원점수 내림차순
# ════════════════════════════════════════════════════════════════════════════
print("\n[STEP 5] 정렬 처리")

LEVEL_SORT = {"최우선 지원": 0, "우선 지원": 1, "모니터링": 2, "안정": 3, "확인 필요": 4}

if "우선지원등급" in df_kr.columns:
    df_kr["_level_sort"] = df_kr["우선지원등급"].map(LEVEL_SORT).fillna(9).astype(int)
    sort_cols  = ["_level_sort", "우선지원점수"]
    sort_asc   = [True, False]
else:
    df_kr["_level_sort"] = 9
    sort_cols  = ["우선지원점수"]
    sort_asc   = [False]

df_kr = df_kr.sort_values(sort_cols, ascending=sort_asc, na_position="last")
df_kr = df_kr.drop(columns=["_level_sort"]).reset_index(drop=True)

print(f"  정렬 완료 - 행 수: {len(df_kr)} (원본 {n_rows_original}행)")
assert len(df_kr) == n_rows_original, "[ERROR] 정렬 후 행 수 변경됨"

# ════════════════════════════════════════════════════════════════════════════
# STEP 6. 변수설명 테이블 생성
# ════════════════════════════════════════════════════════════════════════════
print("\n[STEP 6] 변수설명 테이블 생성")

VAR_META = [
    ("학교코드",           "school_code",                  "학교 식별을 위한 고유 코드",                                       "기본 정보"),
    ("학교명",             "school_name",                  "학교 이름",                                                        "기본 정보"),
    ("시도",               "sido",                         "학교 소재 시도",                                                   "기본 정보"),
    ("시군구",             "sigungu",                      "학교 소재 시군구",                                                 "기본 정보"),
    ("상담인력 공급 점수", "counseling_staff_supply_score","전문상담교사 수와 상담교사 1인당 학생 수를 기준으로 산출한 공급 점수","CSI 하위 변수"),
    ("Wee클래스 운영 점수","wee_class_score",              "Wee클래스 운영 여부를 기준으로 산출한 공급 점수",                  "CSI 하위 변수"),
    ("Wee센터 접근성 점수","wee_center_access_score",      "가장 가까운 Wee센터까지의 직선거리 기준 접근성 점수",              "CSI 하위 변수"),
    ("상담공급지수",        "CSI",                          "상담인력 공급 점수, Wee클래스 운영 점수, Wee센터 접근성 점수의 평균","최종 지수"),
    ("상담수요 규모 점수", "demand_size_score",            "총 학생 수를 기준으로 산출한 상담 수요 규모 점수",                  "CDI 하위 변수"),
    ("실제 상담 이용 점수","counseling_use_score",         "최근 3개년 학생·학부모 상담 이용 자료를 정규화하여 산출한 점수",    "CDI 하위 변수"),
    ("학교폭력 위험 점수", "school_violence_risk_score",   "학교폭력 피해 응답률 기반 위험 점수",                              "CDI 하위 변수"),
    ("상담수요지수",        "CDI",                          "상담수요 규모 점수, 실제 상담 이용 점수, 학교폭력 위험 점수의 평균", "최종 지수"),
    ("우선지원점수",        "priority_score",               "상담수요지수에서 상담공급지수를 뺀 값 (CDI - CSI)",                "최종 지수"),
    ("우선지원등급",        "priority_level",               "우선지원점수를 분위수 기반으로 나눈 정책 검토 등급 (상대비교)",     "최종 지수"),
    ("고정기준 우선지원등급","priority_level_fixed",        "우선지원점수를 고정 기준으로 나눈 보조 등급",                      "최종 지수"),
    ("우선지원점수 결측 사유","priority_missing_reason",   "우선지원점수를 계산하지 못한 이유",                                "보조 정보"),
]

meta_df = pd.DataFrame(VAR_META, columns=["한글 변수명", "원래 변수명", "설명", "지수 구분"])
# 입력 파일에 없는 변수 표시
meta_df["입력파일 존재"] = meta_df["원래 변수명"].apply(
    lambda x: "O" if x in df.columns else "X (누락)"
)

print(f"  변수설명 항목: {len(meta_df)}개")

# ════════════════════════════════════════════════════════════════════════════
# STEP 7. 엑셀 파일 생성 및 서식 적용 (openpyxl)
# ════════════════════════════════════════════════════════════════════════════
print("\n[STEP 7] 엑셀 파일 생성 및 서식 적용")

wb = Workbook()

# ── 공통 스타일 정의 ──────────────────────────────────────────────────────
HEADER_FONT   = Font(name="맑은 고딕", bold=True, size=10, color="FFFFFF")
HEADER_FILL   = PatternFill("solid", fgColor="2E5FA3")   # 진한 파랑
SUB_HDR_FILL  = PatternFill("solid", fgColor="4472C4")   # 연한 파랑 (변수설명 헤더)
BODY_FONT     = Font(name="맑은 고딕", size=10)
CENTER_ALIGN  = Alignment(horizontal="center", vertical="center", wrap_text=False)
LEFT_ALIGN    = Alignment(horizontal="left",   vertical="center", wrap_text=False)
THIN_BORDER   = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)

# 등급별 행 배경색
LEVEL_FILL = {
    "최우선 지원": PatternFill("solid", fgColor="FFE0E0"),   # 연한 빨강
    "우선 지원":   PatternFill("solid", fgColor="FFF2CC"),   # 연한 노랑
    "모니터링":    PatternFill("solid", fgColor="EBF3FB"),   # 연한 파랑
    "안정":        PatternFill("solid", fgColor="F2F2F2"),   # 연한 회색
}

def set_col_width(ws, col_idx, content_list, header_str, min_w=8, max_w=30):
    """열 내용 기준 자동 너비 설정"""
    max_len = max(
        len(str(header_str)),
        max((len(str(v)) for v in content_list if v is not None), default=0),
    )
    ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, min_w), max_w)

# ── 시트 1: 최종요약표 ────────────────────────────────────────────────────
ws1 = wb.active
ws1.title = "최종요약표"

# 점수형 열 인덱스 파악 (1-based)
score_col_indices = set()
for c_idx, col_name in enumerate(df_kr.columns, start=1):
    if col_name in SCORE_COLS_KR:
        score_col_indices.add(c_idx)

# 우선지원등급 열 인덱스
level_col_idx = None
for c_idx, col_name in enumerate(df_kr.columns, start=1):
    if col_name == "우선지원등급":
        level_col_idx = c_idx

# 헤더 행 작성
for c_idx, col_name in enumerate(df_kr.columns, start=1):
    cell = ws1.cell(row=1, column=c_idx, value=col_name)
    cell.font      = HEADER_FONT
    cell.fill      = HEADER_FILL
    cell.alignment = CENTER_ALIGN
    cell.border    = THIN_BORDER

# 데이터 행 작성
for r_idx, row in enumerate(df_kr.itertuples(index=False), start=2):
    # 이 행의 우선지원등급 값 (행 배경색용)
    level_val = getattr(row, "우선지원등급", None) if "우선지원등급" in df_kr.columns else None
    row_fill  = LEVEL_FILL.get(level_val, None)

    for c_idx, val in enumerate(row, start=1):
        # NaN 처리
        if isinstance(val, float) and np.isnan(val):
            val = None

        cell = ws1.cell(row=r_idx, column=c_idx, value=val)
        cell.font   = BODY_FONT
        cell.border = THIN_BORDER

        if c_idx in score_col_indices:
            cell.number_format = "0.000"
            cell.alignment     = CENTER_ALIGN
        elif c_idx == 1:   # 학교코드
            cell.number_format = "@"
            cell.alignment     = CENTER_ALIGN
        else:
            cell.alignment = LEFT_ALIGN

        if row_fill:
            cell.fill = row_fill

# 첫 행 고정
ws1.freeze_panes = "A2"

# 열 너비 자동 조정
for c_idx, col_name in enumerate(df_kr.columns, start=1):
    col_vals = df_kr.iloc[:, c_idx - 1].tolist()
    set_col_width(ws1, c_idx, col_vals, col_name)

print(f"  '최종요약표' 시트 작성 완료 ({len(df_kr)}행 × {len(df_kr.columns)}열)")

# ── 시트 2: 변수설명 ─────────────────────────────────────────────────────
ws2 = wb.create_sheet("변수설명")

for c_idx, col_name in enumerate(meta_df.columns, start=1):
    cell = ws2.cell(row=1, column=c_idx, value=col_name)
    cell.font      = Font(name="맑은 고딕", bold=True, size=10, color="FFFFFF")
    cell.fill      = SUB_HDR_FILL
    cell.alignment = CENTER_ALIGN
    cell.border    = THIN_BORDER

for r_idx, row in enumerate(meta_df.itertuples(index=False), start=2):
    for c_idx, val in enumerate(row, start=1):
        cell = ws2.cell(row=r_idx, column=c_idx, value=val)
        cell.font      = BODY_FONT
        cell.alignment = LEFT_ALIGN
        cell.border    = THIN_BORDER
        # 지수 구분별 배경색
        grp = row[3]  # 지수 구분 열
        if grp == "CSI 하위 변수":
            cell.fill = PatternFill("solid", fgColor="DDEEFF")
        elif grp == "CDI 하위 변수":
            cell.fill = PatternFill("solid", fgColor="DDFFDD")
        elif grp == "최종 지수":
            cell.fill = PatternFill("solid", fgColor="FFF5CC")

ws2.freeze_panes = "A2"

for c_idx, col_name in enumerate(meta_df.columns, start=1):
    col_vals = meta_df.iloc[:, c_idx - 1].tolist()
    set_col_width(ws2, c_idx, col_vals, col_name, max_w=60)

print(f"  '변수설명' 시트 작성 완료 ({len(meta_df)}행 × {len(meta_df.columns)}열)")

# ════════════════════════════════════════════════════════════════════════════
# STEP 8. 저장 및 검증
# ════════════════════════════════════════════════════════════════════════════
print("\n[STEP 8] 저장 및 검증")

wb.save(OUTPUT_PATH)
print(f"  저장 완료: {OUTPUT_PATH}")

# 검증: 행 수·열 수 확인
verify = pd.read_excel(OUTPUT_PATH, sheet_name="최종요약표")
assert len(verify) == n_rows_original, f"[ERROR] 저장 후 행 수 불일치: {len(verify)} ≠ {n_rows_original}"
print(f"  검증 완료: {len(verify)}행 × {len(verify.columns)}열 (원본 {n_rows_original}행 일치)")
print(f"  원본 파일 보존 확인: {INPUT_PATH.exists()} (수정 없음)")

# ════════════════════════════════════════════════════════════════════════════
# 최종 요약 출력
# ════════════════════════════════════════════════════════════════════════════
print("\n" + "=" * 60)
print("한글 표시명 우선지원 요약 파일 생성 완료")
print("=" * 60)
print(f"\n출력 파일: {OUTPUT_PATH}")
print(f"  시트: 최종요약표 ({len(df_kr)}행 × {len(df_kr.columns)}열)")
print(f"  시트: 변수설명   ({len(meta_df)}행 × {len(meta_df.columns)}열)")

if missing_vars:
    print(f"\n[참고] 입력 파일에 없어 변환에서 제외된 변수 ({len(missing_vars)}개):")
    for v in missing_vars:
        print(f"  - {v} → {COL_MAP[v]}")

print("\n정렬 기준: 우선지원등급 우선(최우선→안정), 우선지원점수 내림차순")
print("서식 적용: 헤더 굵게, 행 배경색(등급별), 점수형 소수점 3자리, 첫 행 고정")
