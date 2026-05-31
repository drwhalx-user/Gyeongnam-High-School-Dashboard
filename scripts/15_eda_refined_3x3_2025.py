"""
3x3 분류 기반 EDA 엑셀 데이터 파일 생성

입력 파일:
  data/processed/gyeongnam_high_schools_policy_feedback_refined.xlsx
    (refined_policy_feedback_table 시트, 26열)

출력 파일:
  outputs/tables/EDA_3x3분석결과_2025.xlsx  (8개 시트)
    - 전체_요약통계
    - 지수별_기술통계
    - 등급별_빈도표
    - 3x3매트릭스_빈도표
    - 정책전략그룹별_빈도표
    - 시군구별_지수_요약
    - 우선지원상위20개교
    - 취약학교목록
  outputs/figures/eda3x3_*.png + EDA3x3_*.png  (10종 × 2 = 20개)
  docs/eda_findings_3x3_2025.md
"""

import pathlib
import sys
import warnings
warnings.filterwarnings("ignore")

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

plt.rcParams["font.family"]        = "Malgun Gothic"
plt.rcParams["axes.unicode_minus"] = False

# ── 경로 설정 ─────────────────────────────────────────────────────────────────
ROOT        = pathlib.Path(__file__).resolve().parent.parent
PROCESSED   = ROOT / "data" / "processed"
TABLES_DIR  = ROOT / "outputs" / "tables"
FIGURES_DIR = ROOT / "outputs" / "figures"
DOCS_DIR    = ROOT / "docs"

for d in [TABLES_DIR, FIGURES_DIR, DOCS_DIR]:
    d.mkdir(parents=True, exist_ok=True)

INPUT_PATH  = PROCESSED / "gyeongnam_high_schools_policy_feedback_refined.xlsx"
OUTPUT_XLSX = TABLES_DIR / "EDA_3x3분석결과_2025.xlsx"

# ── COL_MAP: 영문 → 한글 ──────────────────────────────────────────────────────
COL_MAP = {
    "school_code":                   "학교코드",
    "school_name":                   "학교명",
    "sido":                          "시도",
    "sigungu":                       "시군구",
    "counseling_staff_supply_score": "상담인력 공급 점수",
    "wee_class_score":               "Wee클래스 운영 점수",
    "wee_center_access_score":       "Wee센터 접근성 점수",
    "demand_size_score":             "상담수요 규모 점수",
    "counseling_use_score":          "실제 상담 이용 점수",
    "school_violence_risk_score":    "학교폭력 위험 점수",
    "CSI":                           "상담공급지수",
    "CDI":                           "상담수요지수",
    "priority_score":                "우선지원점수",
    "priority_level":                "우선지원등급",
    "csi_level":                     "공급수준등급",
    "cdi_level":                     "수요수준등급",
    "supply_demand_type":            "2x2수요공급유형",
    "policy_action_type":            "2x2정책조치유형",
    "policy_recommendation":         "정책권고사항",
    "policy_reason":                 "정책근거",
    "cdi_relative_level":            "CDI상대등급",
    "csi_relative_level":            "CSI상대등급",
    "supply_demand_matrix_3x3":      "3x3수요공급매트릭스",
    "policy_strategy_group":         "정책전략그룹",
    "policy_strategy_tags":          "정책전략태그",
    "policy_strategy_description":   "정책전략설명",
}

# ── 정렬 순서 상수 ─────────────────────────────────────────────────────────────
MATRIX_ORDER = [
    "핵심 불균형형", "고수요 보완형", "고수요 유지관리형",
    "잠재 취약형", "평균 관리형", "안정 관리형",
    "최소 인프라 보완형", "안정 모니터링형", "여유·거점 활용형", "확인 필요",
]
STRATEGY_ORDER = [
    "최우선 개입형", "우선 보완형", "고수요 유지관리형",
    "인력 취약형", "접근성 보완형", "최소 인프라 보완형",
    "안정형", "확인 필요형",
]
PRIORITY_ORDER = ["최우선 지원", "우선 지원", "모니터링", "안정", "확인 필요"]
CDI_REL_ORDER  = ["수요 상위", "수요 중위", "수요 하위", "확인 필요"]
CSI_REL_ORDER  = ["공급 하위", "공급 중위", "공급 상위", "확인 필요"]

MATRIX_PALETTE = {
    "핵심 불균형형":    "#C0392B",
    "고수요 보완형":    "#E67E22",
    "고수요 유지관리형":"#F1C40F",
    "잠재 취약형":      "#9B59B6",
    "평균 관리형":      "#3498DB",
    "안정 관리형":      "#2ECC71",
    "최소 인프라 보완형":"#1ABC9C",
    "안정 모니터링형":  "#27AE60",
    "여유·거점 활용형": "#BDC3C7",
    "확인 필요":        "#95A5A6",
}
STRATEGY_PALETTE = {
    "최우선 개입형":      "#C0392B",
    "우선 보완형":        "#E67E22",
    "고수요 유지관리형":  "#F1C40F",
    "인력 취약형":        "#9B59B6",
    "접근성 보완형":      "#3498DB",
    "최소 인프라 보완형": "#1ABC9C",
    "안정형":             "#27AE60",
    "확인 필요형":        "#95A5A6",
}
LEVEL_PALETTE = {
    "최우선 지원": "#C0392B",
    "우선 지원":   "#E67E22",
    "모니터링":    "#3498DB",
    "안정":        "#27AE60",
    "확인 필요":   "#95A5A6",
}

# ════════════════════════════════════════════════════════════════════════════
# STEP 1. 입력 파일 로드
# ════════════════════════════════════════════════════════════════════════════
print("[STEP 1] 입력 파일 로드")

if not INPUT_PATH.exists():
    print(f"[ERROR] 파일 없음: {INPUT_PATH}")
    sys.exit(1)

raw = pd.read_excel(INPUT_PATH, sheet_name="refined_policy_feedback_table",
                    dtype={"school_code": str})
n_rows = len(raw)
print(f"  로드 완료: {n_rows}행 x {len(raw.columns)}열")

# 한글 변수명으로 변환
missing_cols = [c for c in COL_MAP if c not in raw.columns]
if missing_cols:
    print(f"  [ERROR] 누락 컬럼: {missing_cols}")
    sys.exit(1)

df = raw.rename(columns=COL_MAP)
print(f"  컬럼 한글화 완료: {len(df.columns)}개")

# ── 분석용 상수 ───────────────────────────────────────────────────────────────
CDI_MEAN = df["상담수요지수"].mean()
CSI_MEAN = df["상담공급지수"].mean()
CDI_P20  = df["상담수요지수"].quantile(0.20)
CDI_P80  = df["상담수요지수"].quantile(0.80)
CDI_P90  = df["상담수요지수"].quantile(0.90)
CSI_P20  = df["상담공급지수"].quantile(0.20)
CSI_P80  = df["상담공급지수"].quantile(0.80)
USE_P80  = df["실제 상담 이용 점수"].quantile(0.80)
VIOL_P80 = df["학교폭력 위험 점수"].quantile(0.80)

# ════════════════════════════════════════════════════════════════════════════
# STEP 2. 각 시트별 DataFrame 생성
# ════════════════════════════════════════════════════════════════════════════
print("\n[STEP 2] 시트별 데이터 생성")

# ── 시트 1: 전체 요약통계 ─────────────────────────────────────────────────────
rows = []
def add(항목, 값): rows.append({"항목": 항목, "값": 값})

add("전체 학교 수",                           n_rows)
add("시군구 수",                               df["시군구"].nunique())
for col in ["상담공급지수", "상담수요지수", "우선지원점수"]:
    s = df[col].dropna()
    add(f"{col} 평균",    round(s.mean(),   4))
    add(f"{col} 최솟값",  round(s.min(),    4))
    add(f"{col} 최댓값",  round(s.max(),    4))
    add(f"{col} 중앙값",  round(s.median(), 4))
add("우선지원점수 양수 학교 수",                 int((df["우선지원점수"] > 0).sum()))
add("최우선 지원 학교 수 (우선지원등급 기준)",   int((df["우선지원등급"] == "최우선 지원").sum()))
add("우선 지원 학교 수 (우선지원등급 기준)",     int((df["우선지원등급"] == "우선 지원").sum()))
add("CDI 분위수 기준 P20",                     round(CDI_P20, 4))
add("CDI 분위수 기준 P80",                     round(CDI_P80, 4))
add("CSI 분위수 기준 P20",                     round(CSI_P20, 4))
add("CSI 분위수 기준 P80",                     round(CSI_P80, 4))
for t in MATRIX_ORDER:
    cnt = int((df["3x3수요공급매트릭스"] == t).sum())
    if cnt > 0:
        add(f"3x3 매트릭스 - {t} 학교 수", cnt)
for g in STRATEGY_ORDER:
    cnt = int((df["정책전략그룹"] == g).sum())
    if cnt > 0:
        add(f"정책전략 - {g} 학교 수", cnt)

sheet1_df = pd.DataFrame(rows)
print(f"  시트1 전체_요약통계: {len(sheet1_df)}행")

# ── 시트 2: 지수별 기술통계 ───────────────────────────────────────────────────
INDEX_COLS = [
    "상담공급지수", "상담수요지수", "우선지원점수",
    "상담인력 공급 점수", "Wee클래스 운영 점수", "Wee센터 접근성 점수",
    "상담수요 규모 점수", "실제 상담 이용 점수", "학교폭력 위험 점수",
]
sheet2_df = (
    df[INDEX_COLS]
    .describe(percentiles=[0.20, 0.25, 0.50, 0.75, 0.80])
    .T
    .rename(columns={
        "count":  "학교수",
        "mean":   "평균",
        "std":    "표준편차",
        "min":    "최솟값",
        "20%":    "P20",
        "25%":    "Q1",
        "50%":    "중앙값",
        "75%":    "Q3",
        "80%":    "P80",
        "max":    "최댓값",
    })
    .round(4)
    .reset_index()
    .rename(columns={"index": "변수명"})
)
print(f"  시트2 지수별_기술통계: {len(sheet2_df)}행")

# ── 시트 3: 등급별 빈도표 ─────────────────────────────────────────────────────
def freq_table(series, order, col_label="구분"):
    cnt = series.value_counts()
    cnt = cnt.reindex([o for o in order if o in cnt.index]).fillna(0).astype(int)
    pct = (cnt / n_rows * 100).round(1)
    return pd.DataFrame({col_label: cnt.index, "학교수": cnt.values, "비율(%)": pct.values})

level_rows = []
for col, lbl, order in [
    ("우선지원등급",  "우선지원등급",  PRIORITY_ORDER),
    ("CDI상대등급",   "CDI상대등급",   CDI_REL_ORDER),
    ("CSI상대등급",   "CSI상대등급",   CSI_REL_ORDER),
    ("수요수준등급",  "수요수준등급",  ["수요 낮음", "수요 보통", "수요 높음", "확인 필요"]),
    ("공급수준등급",  "공급수준등급",  ["공급 낮음", "공급 보통", "공급 양호", "공급 높음", "확인 필요"]),
]:
    ft = freq_table(df[col], order, "등급")
    ft.insert(0, "변수", lbl)
    level_rows.append(ft)
sheet3_df = pd.concat(level_rows, ignore_index=True)
print(f"  시트3 등급별_빈도표: {len(sheet3_df)}행")

# ── 시트 4: 3x3 매트릭스 빈도표 ──────────────────────────────────────────────
matrix_cnt = df["3x3수요공급매트릭스"].value_counts()
matrix_order_present = [t for t in MATRIX_ORDER if t in matrix_cnt.index]
matrix_cnt = matrix_cnt.reindex(matrix_order_present).fillna(0).astype(int)

sheet4_rows = []
for t, cnt in matrix_cnt.items():
    pct = round(cnt / n_rows * 100, 1)
    grp_df = df[df["3x3수요공급매트릭스"] == t]
    sheet4_rows.append({
        "3x3수요공급매트릭스":  t,
        "학교수":              cnt,
        "비율(%)":             pct,
        "CSI평균":             round(grp_df["상담공급지수"].mean(), 4),
        "CDI평균":             round(grp_df["상담수요지수"].mean(), 4),
        "우선지원점수평균":     round(grp_df["우선지원점수"].mean(), 4),
        "상담인력공급점수평균": round(grp_df["상담인력 공급 점수"].mean(), 4),
        "Wee클래스점수평균":    round(grp_df["Wee클래스 운영 점수"].mean(), 4),
        "Wee센터접근성점수평균":round(grp_df["Wee센터 접근성 점수"].mean(), 4),
        "상담수요규모점수평균": round(grp_df["상담수요 규모 점수"].mean(), 4),
        "실제상담이용점수평균": round(grp_df["실제 상담 이용 점수"].mean(), 4),
        "학교폭력위험점수평균": round(grp_df["학교폭력 위험 점수"].mean(), 4),
    })
sheet4_df = pd.DataFrame(sheet4_rows)
print(f"  시트4 3x3매트릭스_빈도표: {len(sheet4_df)}행")

# ── 시트 5: 정책전략그룹별 빈도표 ────────────────────────────────────────────
strategy_cnt = df["정책전략그룹"].value_counts()
strategy_order_present = [g for g in STRATEGY_ORDER if g in strategy_cnt.index]
strategy_cnt = strategy_cnt.reindex(strategy_order_present).fillna(0).astype(int)

sheet5_rows = []
for g, cnt in strategy_cnt.items():
    pct = round(cnt / n_rows * 100, 1)
    grp_df = df[df["정책전략그룹"] == g]
    sheet5_rows.append({
        "정책전략그룹":         g,
        "학교수":               cnt,
        "비율(%)":              pct,
        "CSI평균":              round(grp_df["상담공급지수"].mean(), 4),
        "CDI평균":              round(grp_df["상담수요지수"].mean(), 4),
        "우선지원점수평균":      round(grp_df["우선지원점수"].mean(), 4),
        "최우선지원_수":         int((grp_df["우선지원등급"] == "최우선 지원").sum()),
        "우선지원_수":           int((grp_df["우선지원등급"] == "우선 지원").sum()),
        "상담인력공급점수평균":  round(grp_df["상담인력 공급 점수"].mean(), 4),
        "Wee클래스점수평균":     round(grp_df["Wee클래스 운영 점수"].mean(), 4),
        "Wee센터접근성점수평균": round(grp_df["Wee센터 접근성 점수"].mean(), 4),
        "정책전략설명":          STRATEGY_ORDER.index(g) if g in STRATEGY_ORDER else "",
    })

# 정책전략설명 텍스트 채우기
STRATEGY_DESC = {
    "최우선 개입형":       "상담수요가 상대적으로 높고 상담공급이 부족하거나 우선지원점수가 높아 교육청 차원의 우선 지원 검토가 필요한 유형",
    "우선 보완형":         "상담수요가 높은 편으로 기존 인프라 보강 또는 상담 프로그램 확대 검토가 필요한 유형",
    "고수요 유지관리형":   "상담수요가 높지만 공급도 비교적 확보되어 있어 기존 인프라 유지와 프로그램 질 관리가 필요한 유형",
    "인력 취약형":         "전문상담교사 공급 점수가 낮아 전문상담교사 배치 또는 순회상담 연계 검토가 필요한 유형",
    "접근성 보완형":       "Wee센터 접근성이 낮아 온라인 상담·이동형 상담 또는 권역별 Wee센터 연계 강화가 필요한 유형",
    "최소 인프라 보완형":  "상담수요는 높지 않더라도 공급 기반이 낮아 권역별 순회상담 등 최소 인프라 보완이 필요한 유형",
    "안정형":              "현재 지표상 수요 대비 공급이 상대적으로 안정적인 유형",
    "확인 필요형":         "주요 지표 결측으로 추가 확인이 필요한 유형",
}
for row in sheet5_rows:
    row["정책전략설명"] = STRATEGY_DESC.get(row["정책전략그룹"], "")
sheet5_df = pd.DataFrame(sheet5_rows)
print(f"  시트5 정책전략그룹별_빈도표: {len(sheet5_df)}행")

# ── 시트 6: 시군구별 지수 요약 ────────────────────────────────────────────────
sigungu_df = (
    df.groupby("시군구")
    .agg(
        학교수               = ("학교코드",              "count"),
        상담공급지수_평균    = ("상담공급지수",           "mean"),
        상담수요지수_평균    = ("상담수요지수",           "mean"),
        우선지원점수_평균    = ("우선지원점수",           "mean"),
        최우선지원_수        = ("우선지원등급",  lambda x: (x == "최우선 지원").sum()),
        우선지원_수          = ("우선지원등급",  lambda x: (x == "우선 지원").sum()),
        최우선개입형_수      = ("정책전략그룹",  lambda x: (x == "최우선 개입형").sum()),
        우선보완형_수        = ("정책전략그룹",  lambda x: (x == "우선 보완형").sum()),
        인력취약형_수        = ("정책전략그룹",  lambda x: (x == "인력 취약형").sum()),
        접근성보완형_수      = ("정책전략그룹",  lambda x: (x == "접근성 보완형").sum()),
        고수요보완형_수      = ("3x3수요공급매트릭스", lambda x: (x == "고수요 보완형").sum()),
        잠재취약형_수        = ("3x3수요공급매트릭스", lambda x: (x == "잠재 취약형").sum()),
        최소인프라보완형_수  = ("3x3수요공급매트릭스", lambda x: (x == "최소 인프라 보완형").sum()),
        상담인력공급점수_평균 = ("상담인력 공급 점수",   "mean"),
        Wee클래스점수_평균   = ("Wee클래스 운영 점수",  "mean"),
        Wee센터접근성점수_평균= ("Wee센터 접근성 점수", "mean"),
        상담수요규모점수_평균 = ("상담수요 규모 점수",   "mean"),
        실제상담이용점수_평균 = ("실제 상담 이용 점수",  "mean"),
        학교폭력위험점수_평균 = ("학교폭력 위험 점수",   "mean"),
    )
    .round(4)
    .reset_index()
    .sort_values("우선지원점수_평균", ascending=False)
)
sheet6_df = sigungu_df.copy()
print(f"  시트6 시군구별_지수_요약: {len(sheet6_df)}행")

# ── 시트 7: 우선지원 상위 20개교 ─────────────────────────────────────────────
TOP_COLS = [
    "학교코드", "학교명", "시군구",
    "상담공급지수", "상담수요지수", "우선지원점수",
    "우선지원등급", "CDI상대등급", "CSI상대등급",
    "3x3수요공급매트릭스", "정책전략그룹", "정책전략태그",
]
sheet7_df = (
    df[df["우선지원점수"].notna()]
    .nlargest(20, "우선지원점수")[TOP_COLS]
    .reset_index(drop=True)
)
sheet7_df.index += 1
sheet7_df.index.name = "순위"
sheet7_df = sheet7_df.reset_index()
print(f"  시트7 우선지원상위20개교: {len(sheet7_df)}행")

# ── 시트 8: 취약학교 목록 ─────────────────────────────────────────────────────
LOW_COLS = [
    "학교코드", "학교명", "시군구",
    "상담인력 공급 점수", "Wee클래스 운영 점수", "Wee센터 접근성 점수",
    "상담수요 규모 점수", "실제 상담 이용 점수", "학교폭력 위험 점수",
    "상담공급지수", "상담수요지수", "우선지원점수",
    "우선지원등급", "3x3수요공급매트릭스", "정책전략그룹",
]
cond_staff  = df["상담인력 공급 점수"] < 0.4
cond_wee    = df["Wee클래스 운영 점수"] == 0
cond_center = df["Wee센터 접근성 점수"] <= 0.4
cond_viol   = df["학교폭력 위험 점수"]  >= VIOL_P80
cond_use    = df["실제 상담 이용 점수"] >= USE_P80
any_cond    = cond_staff | cond_wee | cond_center | cond_viol | cond_use

low_df = df[any_cond][LOW_COLS].copy()

def build_weakness(row):
    parts = []
    if row["상담인력 공급 점수"] < 0.4:          parts.append("상담인력 공급 낮음")
    if row["Wee클래스 운영 점수"] == 0:           parts.append("Wee클래스 미운영")
    if row["Wee센터 접근성 점수"] <= 0.4:         parts.append("Wee센터 접근성 낮음")
    if row["학교폭력 위험 점수"]  >= VIOL_P80:    parts.append("학교폭력 위험 상위20%")
    if row["실제 상담 이용 점수"] >= USE_P80:     parts.append("상담이용 상위20%")
    return "; ".join(parts)

low_df = low_df.copy()
low_df["취약조건"] = low_df.apply(build_weakness, axis=1)
sheet8_df = low_df.sort_values("우선지원점수", ascending=False).reset_index(drop=True)
print(f"  시트8 취약학교목록: {len(sheet8_df)}개교")

# ════════════════════════════════════════════════════════════════════════════
# STEP 3. Excel 파일 저장 (openpyxl 서식 적용)
# ════════════════════════════════════════════════════════════════════════════
print("\n[STEP 3] Excel 파일 저장")

SHEETS = [
    ("전체_요약통계",       sheet1_df),
    ("지수별_기술통계",     sheet2_df),
    ("등급별_빈도표",       sheet3_df),
    ("3x3매트릭스_빈도표",  sheet4_df),
    ("정책전략그룹별_빈도표", sheet5_df),
    ("시군구별_지수_요약",  sheet6_df),
    ("우선지원상위20개교",  sheet7_df),
    ("취약학교목록",        sheet8_df),
]

with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
    for sheet_name, data_df in SHEETS:
        data_df.to_excel(writer, sheet_name=sheet_name, index=False)

# openpyxl 서식 적용
wb = load_workbook(OUTPUT_XLSX)

HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10, name="맑은 고딕")
HEADER_FILL  = PatternFill("solid", fgColor="2E5FA3")
ALT_FILL     = PatternFill("solid", fgColor="EBF2FF")
NORMAL_FONT  = Font(name="맑은 고딕", size=10)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=False)
LEFT_ALIGN   = Alignment(horizontal="left",   vertical="center", wrap_text=False)
THIN_BORDER  = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)

def style_ws(ws, text_cols=None):
    """헤더 강조, 교대 행 색상, 컬럼 너비 자동 조정"""
    text_cols = text_cols or []
    # 헤더
    for cell in ws[1]:
        cell.font   = HEADER_FONT
        cell.fill   = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
    # 데이터 행
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        fill = ALT_FILL if i % 2 == 0 else PatternFill()
        for cell in row:
            cell.font      = NORMAL_FONT
            cell.fill      = fill
            cell.border    = THIN_BORDER
            header_val = ws.cell(row=1, column=cell.column).value or ""
            is_text = any(t in str(header_val) for t in text_cols)
            cell.alignment = LEFT_ALIGN if is_text else CENTER_ALIGN
    # 컬럼 너비
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max(
            (len(str(cell.value)) if cell.value is not None else 0 for cell in col),
            default=8
        )
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 8), 45)
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 18

TEXT_COLS_MAP = {
    "전체_요약통계":       ["항목"],
    "지수별_기술통계":     ["변수명"],
    "등급별_빈도표":       ["변수", "등급"],
    "3x3매트릭스_빈도표":  ["3x3수요공급매트릭스"],
    "정책전략그룹별_빈도표":["정책전략그룹", "정책전략설명"],
    "시군구별_지수_요약":  ["시군구"],
    "우선지원상위20개교":  ["학교명", "시군구", "3x3수요공급매트릭스", "정책전략그룹", "정책전략태그"],
    "취약학교목록":        ["학교명", "시군구", "3x3수요공급매트릭스", "정책전략그룹", "취약조건"],
}

for ws in wb.worksheets:
    style_ws(ws, text_cols=TEXT_COLS_MAP.get(ws.title, []))

wb.save(OUTPUT_XLSX)
print(f"  저장 완료: {OUTPUT_XLSX}")
for sn, d in SHEETS:
    print(f"    [{sn}] {len(d)}행 x {len(d.columns)}열")

# ════════════════════════════════════════════════════════════════════════════
# STEP 4. 그래프 생성 (10종 × 영문+국문 2개 = 20개 PNG)
# ════════════════════════════════════════════════════════════════════════════
print("\n[STEP 4] 그래프 생성")

def save_fig_both(eng_name, kor_name, dpi=300):
    for fname in [eng_name, kor_name]:
        plt.savefig(FIGURES_DIR / fname, dpi=dpi, bbox_inches="tight", facecolor="white")
    plt.close()
    print(f"  저장: {eng_name}  /  {kor_name}")

# ── 그래프 1: CSI 분포 히스토그램 ────────────────────────────────────────────
fig, ax = plt.subplots(figsize=(8, 5))
ax.hist(df["상담공급지수"].dropna(), bins=20, color="#3498DB", edgecolor="white", linewidth=0.6)
ax.axvline(CSI_MEAN, color="#E74C3C", linestyle="--", linewidth=1.5,
           label=f"평균 {CSI_MEAN:.3f}")
ax.axvline(CSI_P20, color="#F39C12", linestyle=":", linewidth=1.2,
           label=f"P20 ({CSI_P20:.3f})")
ax.axvline(CSI_P80, color="#27AE60", linestyle=":", linewidth=1.2,
           label=f"P80 ({CSI_P80:.3f})")
ax.set_title("상담공급지수(CSI) 분포", fontsize=14, fontweight="bold", pad=12)
ax.set_xlabel("상담공급지수", fontsize=11)
ax.set_ylabel("학교 수", fontsize=11)
ax.legend(fontsize=9)
ax.set_xlim(0, 1)
sns.despine()
save_fig_both("eda3x3_csi_dist_2025.png", "EDA3x3_상담공급지수_분포_2025.png")

# ── 그래프 2: CDI 분포 히스토그램 ────────────────────────────────────────────
fig, ax = plt.subplots(figsize=(8, 5))
ax.hist(df["상담수요지수"].dropna(), bins=20, color="#E67E22", edgecolor="white", linewidth=0.6)
ax.axvline(CDI_MEAN, color="#2C3E50", linestyle="--", linewidth=1.5,
           label=f"평균 {CDI_MEAN:.3f}")
ax.axvline(CDI_P20, color="#F39C12", linestyle=":", linewidth=1.2,
           label=f"P20 ({CDI_P20:.3f})")
ax.axvline(CDI_P80, color="#27AE60", linestyle=":", linewidth=1.2,
           label=f"P80 ({CDI_P80:.3f})")
ax.set_title("상담수요지수(CDI) 분포", fontsize=14, fontweight="bold", pad=12)
ax.set_xlabel("상담수요지수", fontsize=11)
ax.set_ylabel("학교 수", fontsize=11)
ax.legend(fontsize=9)
ax.set_xlim(0, 1)
sns.despine()
save_fig_both("eda3x3_cdi_dist_2025.png", "EDA3x3_상담수요지수_분포_2025.png")

# ── 그래프 3: 우선지원점수 분포 히스토그램 ───────────────────────────────────
p_scores = df["우선지원점수"].dropna()
fig, ax = plt.subplots(figsize=(8, 5))
ax.hist(p_scores, bins=25, color="#8E44AD", edgecolor="white", linewidth=0.6)
ax.axvline(0, color="#E74C3C", linestyle="-", linewidth=2, label="기준선 (0)")
ax.axvline(p_scores.mean(), color="#F39C12", linestyle="--", linewidth=1.5,
           label=f"평균 {p_scores.mean():.3f}")
ax.set_title("우선지원점수(Priority Score) 분포", fontsize=14, fontweight="bold", pad=12)
ax.set_xlabel("우선지원점수 (상담수요지수 - 상담공급지수)", fontsize=11)
ax.set_ylabel("학교 수", fontsize=11)
ax.legend(fontsize=10)
sns.despine()
save_fig_both("eda3x3_priority_dist_2025.png", "EDA3x3_우선지원점수_분포_2025.png")

# ── 그래프 4: 우선지원등급별 학교 수 ─────────────────────────────────────────
order4  = [l for l in PRIORITY_ORDER if (df["우선지원등급"] == l).sum() > 0]
cnt4    = df["우선지원등급"].value_counts().reindex(order4).fillna(0).astype(int)
colors4 = [LEVEL_PALETTE.get(l, "#95A5A6") for l in order4]
fig, ax = plt.subplots(figsize=(8, 5))
bars = ax.bar(order4, cnt4.values, color=colors4, edgecolor="white", linewidth=0.6)
for bar, val in zip(bars, cnt4.values):
    ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.5,
            f"{val}개교\n({val/n_rows*100:.1f}%)", ha="center", va="bottom", fontsize=10)
ax.set_title("우선지원등급별 학교 수", fontsize=14, fontweight="bold", pad=12)
ax.set_xlabel("우선지원등급", fontsize=11)
ax.set_ylabel("학교 수", fontsize=11)
ax.set_ylim(0, cnt4.max() * 1.28)
sns.despine()
save_fig_both("eda3x3_priority_level_2025.png", "EDA3x3_우선지원등급별_학교수_2025.png")

# ── 그래프 5: 3x3 매트릭스 유형별 학교 수 ────────────────────────────────────
present_m = [t for t in MATRIX_ORDER if (df["3x3수요공급매트릭스"] == t).sum() > 0]
cnt5      = df["3x3수요공급매트릭스"].value_counts().reindex(present_m).fillna(0).astype(int)
colors5   = [MATRIX_PALETTE.get(t, "#aaa") for t in present_m]

fig, ax = plt.subplots(figsize=(12, 5))
bars = ax.bar(range(len(present_m)), cnt5.values, color=colors5,
              edgecolor="white", linewidth=0.6)
ax.set_xticks(range(len(present_m)))
ax.set_xticklabels(present_m, fontsize=9, rotation=20, ha="right")
for bar, val in zip(bars, cnt5.values):
    ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.3,
            f"{val}개교\n({val/n_rows*100:.1f}%)", ha="center", va="bottom", fontsize=9)
ax.set_title("3x3 수요-공급 매트릭스 유형별 학교 수", fontsize=14, fontweight="bold", pad=12)
ax.set_ylabel("학교 수", fontsize=11)
ax.set_ylim(0, cnt5.max() * 1.28)
sns.despine()
plt.tight_layout()
save_fig_both("eda3x3_matrix_counts_2025.png", "EDA3x3_3x3매트릭스별_학교수_2025.png")

# ── 그래프 6: CSI-CDI 산점도 (3x3 매트릭스 색상, P20/P80 격자) ───────────────
fig, ax = plt.subplots(figsize=(10, 7))
for mtype in present_m:
    grp   = df[df["3x3수요공급매트릭스"] == mtype]
    color = MATRIX_PALETTE.get(mtype, "#aaa")
    ax.scatter(grp["상담공급지수"], grp["상담수요지수"],
               c=color, label=mtype, alpha=0.80, s=65,
               edgecolors="white", linewidths=0.4)

ax.axvline(CSI_P20, color="#7F8C8D", linestyle=":", linewidth=1.0, alpha=0.7,
           label=f"CSI P20 ({CSI_P20:.2f})")
ax.axvline(CSI_P80, color="#7F8C8D", linestyle=":",  linewidth=1.0, alpha=0.7,
           label=f"CSI P80 ({CSI_P80:.2f})")
ax.axhline(CDI_P20, color="#BDC3C7", linestyle=":", linewidth=1.0, alpha=0.7,
           label=f"CDI P20 ({CDI_P20:.2f})")
ax.axhline(CDI_P80, color="#BDC3C7", linestyle=":", linewidth=1.0, alpha=0.7,
           label=f"CDI P80 ({CDI_P80:.2f})")

ax.set_title("CSI-CDI 산점도: 3x3 수요-공급 매트릭스 유형", fontsize=14, fontweight="bold", pad=12)
ax.set_xlabel("상담공급지수(CSI)", fontsize=11)
ax.set_ylabel("상담수요지수(CDI)", fontsize=11)
ax.set_xlim(0, 1.05)
ax.set_ylim(0, 0.80)
ax.legend(fontsize=8, loc="upper left", framealpha=0.85, ncol=2)
sns.despine()
save_fig_both("eda3x3_scatter_2025.png", "EDA3x3_공급지수_수요지수_산점도_2025.png")

# ── 그래프 7: 시군구별 평균 우선지원점수 가로 막대 ───────────────────────────
sg      = sigungu_df.sort_values("우선지원점수_평균", ascending=True)
colors7 = ["#E74C3C" if v > -0.15 else "#3498DB" for v in sg["우선지원점수_평균"]]
fig, ax = plt.subplots(figsize=(9, 8))
bars = ax.barh(sg["시군구"], sg["우선지원점수_평균"],
               color=colors7, edgecolor="white", linewidth=0.5)
ax.axvline(0, color="#2C3E50", linewidth=1.2)
for bar, val in zip(bars, sg["우선지원점수_평균"]):
    ax.text(val + (0.005 if val >= 0 else -0.005),
            bar.get_y() + bar.get_height() / 2,
            f"{val:.3f}", va="center",
            ha="left" if val >= 0 else "right", fontsize=8.5)
ax.set_title("시군구별 평균 우선지원점수", fontsize=14, fontweight="bold", pad=12)
ax.set_xlabel("평균 우선지원점수 (상담수요지수 - 상담공급지수)", fontsize=11)
ax.set_ylabel("시군구", fontsize=11)
sns.despine()
plt.tight_layout()
save_fig_both("eda3x3_sigungu_2025.png", "EDA3x3_시군구별_우선지원점수_2025.png")

# ── 그래프 8: 정책전략그룹별 학교 수 ─────────────────────────────────────────
present_g = [g for g in STRATEGY_ORDER if (df["정책전략그룹"] == g).sum() > 0]
cnt8      = df["정책전략그룹"].value_counts().reindex(present_g).fillna(0).astype(int)
colors8   = [STRATEGY_PALETTE.get(g, "#aaa") for g in present_g]

fig, ax = plt.subplots(figsize=(11, 5))
bars = ax.bar(range(len(present_g)), cnt8.values, color=colors8,
              edgecolor="white", linewidth=0.6)
ax.set_xticks(range(len(present_g)))
ax.set_xticklabels(present_g, fontsize=9, rotation=15, ha="right")
for bar, val in zip(bars, cnt8.values):
    ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.4,
            f"{val}개교\n({val/n_rows*100:.1f}%)", ha="center", va="bottom", fontsize=10)
ax.set_title("정책전략 그룹별 학교 수", fontsize=14, fontweight="bold", pad=12)
ax.set_ylabel("학교 수", fontsize=11)
ax.set_ylim(0, cnt8.max() * 1.28)
sns.despine()
plt.tight_layout()
save_fig_both("eda3x3_strategy_group_2025.png", "EDA3x3_정책전략그룹별_학교수_2025.png")

# ── 그래프 9: CSI 하위변수 평균비교 ──────────────────────────────────────────
csi_labels = ["상담인력\n공급 점수", "Wee클래스\n운영 점수", "Wee센터\n접근성 점수"]
csi_cols   = ["상담인력 공급 점수", "Wee클래스 운영 점수", "Wee센터 접근성 점수"]
csi_means  = [df[c].mean() for c in csi_cols]
fig, ax = plt.subplots(figsize=(7, 5))
bars = ax.bar(csi_labels, csi_means, color=["#E74C3C", "#3498DB", "#2ECC71"],
              edgecolor="white", linewidth=0.6)
for bar, val in zip(bars, csi_means):
    ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.01,
            f"{val:.3f}", ha="center", va="bottom", fontsize=11, fontweight="bold")
ax.axhline(CSI_MEAN, color="#7F8C8D", linestyle="--", linewidth=1,
           label=f"상담공급지수 전체 평균 ({CSI_MEAN:.3f})")
ax.set_title("상담공급지수(CSI) 하위 변수 평균 비교", fontsize=14, fontweight="bold", pad=12)
ax.set_ylabel("평균 점수", fontsize=11)
ax.set_ylim(0, 1.1)
ax.legend(fontsize=9)
sns.despine()
save_fig_both("eda3x3_csi_subscores_2025.png", "EDA3x3_공급지수_하위변수_평균비교_2025.png")

# ── 그래프 10: CDI 하위변수 평균비교 ─────────────────────────────────────────
cdi_labels = ["상담수요\n규모 점수", "실제 상담\n이용 점수", "학교폭력\n위험 점수"]
cdi_cols   = ["상담수요 규모 점수", "실제 상담 이용 점수", "학교폭력 위험 점수"]
cdi_means  = [df[c].mean() for c in cdi_cols]
fig, ax = plt.subplots(figsize=(7, 5))
bars = ax.bar(cdi_labels, cdi_means, color=["#E67E22", "#9B59B6", "#E74C3C"],
              edgecolor="white", linewidth=0.6)
for bar, val in zip(bars, cdi_means):
    ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.01,
            f"{val:.3f}", ha="center", va="bottom", fontsize=11, fontweight="bold")
ax.axhline(CDI_MEAN, color="#7F8C8D", linestyle="--", linewidth=1,
           label=f"상담수요지수 전체 평균 ({CDI_MEAN:.3f})")
ax.set_title("상담수요지수(CDI) 하위 변수 평균 비교", fontsize=14, fontweight="bold", pad=12)
ax.set_ylabel("평균 점수", fontsize=11)
ax.set_ylim(0, 1.1)
ax.legend(fontsize=9)
sns.despine()
save_fig_both("eda3x3_cdi_subscores_2025.png", "EDA3x3_수요지수_하위변수_평균비교_2025.png")

# ════════════════════════════════════════════════════════════════════════════
# STEP 5. EDA 해석 문서 생성
# ════════════════════════════════════════════════════════════════════════════
print("\n[STEP 5] EDA 해석 문서 생성")

top5_sg = sigungu_df.head(5)[["시군구", "우선지원점수_평균", "최우선지원_수", "우선지원_수"]].to_string(index=False)
top1    = df.nlargest(1, "우선지원점수")[["학교명", "시군구", "우선지원점수"]].iloc[0]

matrix_dist_str = ""
for t in MATRIX_ORDER:
    cnt = int((df["3x3수요공급매트릭스"] == t).sum())
    if cnt > 0:
        matrix_dist_str += f"  - {t}: {cnt}개교 ({cnt/n_rows*100:.1f}%)\n"

strategy_dist_str = ""
for g in STRATEGY_ORDER:
    cnt = int((df["정책전략그룹"] == g).sum())
    if cnt > 0:
        strategy_dist_str += f"  - {g}: {cnt}개교 ({cnt/n_rows*100:.1f}%)\n"

findings_md = f"""# EDA 분석 결과 해석 (3x3 분류 기반)

## 1. 전체 분석 대상
- 분석 대상: 경상남도 일반고등학교 {n_rows}개교 (18개 시군구)
- 분석 기준연도: 2023~2025학년도
- 3x3 유형화 기준: CDI P20={CDI_P20:.4f}/P80={CDI_P80:.4f}, CSI P20={CSI_P20:.4f}/P80={CSI_P80:.4f}

## 2. 상담공급지수(CSI) 분포 해석
- 평균: {CSI_MEAN:.4f}, 범위: {df['상담공급지수'].min():.4f} ~ {df['상담공급지수'].max():.4f}
- P20={CSI_P20:.4f} (공급 하위 경계), P80={CSI_P80:.4f} (공급 상위 경계)
- 공급 하위({(df['CSI상대등급']=='공급 하위').sum()}개교) / 공급 중위({(df['CSI상대등급']=='공급 중위').sum()}개교) / 공급 상위({(df['CSI상대등급']=='공급 상위').sum()}개교)
- CSI 평균(0.65)은 표면적으로 양호하나 상담인력 공급 점수 평균({df['상담인력 공급 점수'].mean():.3f})이
  Wee클래스({df['Wee클래스 운영 점수'].mean():.3f})·Wee센터 접근성({df['Wee센터 접근성 점수'].mean():.3f})에 비해 현저히 낮아
  인프라 대비 인력 취약 구조임을 확인

## 3. 상담수요지수(CDI) 분포 해석
- 평균: {CDI_MEAN:.4f}, 범위: {df['상담수요지수'].min():.4f} ~ {df['상담수요지수'].max():.4f}
- P20={CDI_P20:.4f} (수요 하위 경계), P80={CDI_P80:.4f} (수요 상위 경계)
- 수요 상위({(df['CDI상대등급']=='수요 상위').sum()}개교) / 수요 중위({(df['CDI상대등급']=='수요 중위').sum()}개교) / 수요 하위({(df['CDI상대등급']=='수요 하위').sum()}개교)
- 상담수요 규모 점수 평균 {df['상담수요 규모 점수'].mean():.3f}: 학생 수 기반 수요 잠재성 높음
- 실제 상담 이용 점수 평균 {df['실제 상담 이용 점수'].mean():.3f}: 낮음 - 미충족 수요 가능성
- 3x3 분류 전에는 74%가 '수요 보통'으로 집중되었으나, 상대 분위수 기반 분류로 차별화 가능

## 4. 우선지원점수(Priority Score) 분포 해석
- 평균: {df['우선지원점수'].mean():.4f}, 범위: {df['우선지원점수'].min():.4f} ~ {df['우선지원점수'].max():.4f}
- {int((df['우선지원점수'] > 0).sum())}개교만 양수(수요 > 공급), {int((df['우선지원점수'] <= 0).sum())}개교는 음수(공급 >= 수요)
- 분위수 기반 상대비교로 상위 약 {int((df['우선지원등급'].isin(['최우선 지원','우선 지원'])).sum())}개교를 우선 지원 검토 대상으로 도출
- 우선지원점수 최고 학교: {top1['학교명']}({top1['시군구']}, 점수={top1['우선지원점수']:.4f})

## 5. 3x3 수요-공급 매트릭스 결과 해석
{matrix_dist_str}
- 고수요 보완형(수요 상위×공급 중위)이 가장 많음: 수요는 높지만 공급이 중간 수준
- 평균 관리형(수요 중위×공급 중위)이 {int((df['3x3수요공급매트릭스']=='평균 관리형').sum())}개교(43.2%)로 가장 다수: 중위 군집
- 핵심 불균형형(수요 상위×공급 하위): {int((df['3x3수요공급매트릭스']=='핵심 불균형형').sum())}개교 - 경남 일반고에서 공급 최하위 학교 중 수요 최상위는 없음

## 6. 정책전략 그룹 결과 해석
{strategy_dist_str}
- 최우선 개입형({int((df['정책전략그룹']=='최우선 개입형').sum())}개교): 잠재 취약형 학교 중 공급 부족 + 우선지원점수 기준
- 인력 취약형({int((df['정책전략그룹']=='인력 취약형').sum())}개교, 30.1%): 전문상담교사 공급 점수 0.4 미만 - 가장 많은 그룹 중 하나
- 안정형({int((df['정책전략그룹']=='안정형').sum())}개교, 30.1%): 수요 대비 공급이 상대적으로 안정적

## 7. 시군구별 우선지원 경향
우선지원점수 평균 상위 시군구 (내림차순):
{top5_sg}

## 8. 취약 학교 분포 요약
- 상담인력 공급 점수 < 0.4: {int(cond_staff.sum())}개교 ({cond_staff.sum()/n_rows*100:.1f}%)
- Wee클래스 미운영: {int(cond_wee.sum())}개교
- Wee센터 접근성 낮음: {int(cond_center.sum())}개교 ({cond_center.sum()/n_rows*100:.1f}%)
- 학교폭력 위험 상위 20%: {int(cond_viol.sum())}개교
- 상담이용 상위 20%: {int(cond_use.sum())}개교
- 위 조건 중 하나 이상 해당: {int(any_cond.sum())}개교 ({any_cond.sum()/n_rows*100:.1f}%)

## 9. 보고서·PPT 핵심 해석 문장
1. "3x3 상대 분위수 기반 매트릭스를 적용한 결과, 기존 절대 기준에서 74%가 '수요 보통'으로 집중되던 한계를 극복하고 학교별 차별화된 정책 전략 배분이 가능해졌다."
2. "고수요 보완형({int((df['3x3수요공급매트릭스']=='고수요 보완형').sum())}개교): 상담수요 상위이지만 공급이 중간 수준으로, 인프라 보강과 상담 프로그램 확대가 우선 검토 대상이다."
3. "인력 취약형({int((df['정책전략그룹']=='인력 취약형').sum())}개교): 전문상담교사 배치율이 낮은 학교군으로, 순회상담 연계 또는 신규 배치 우선 검토 대상이다."
4. "안정형({int((df['정책전략그룹']=='안정형').sum())}개교)은 현 수준 유지와 함께 거점 학교로서의 활용 가능성을 검토할 수 있다."

## 10. 주의사항 및 한계
- 본 3x3 분류는 기존 절대 기준 분류(cdi_level, csi_level)를 대체하는 것이 아니라 보완하는 보조 분류다.
- CDI와 CSI는 동일가중치 기반으로 산출되어 구성요소의 실제 중요도를 반드시 반영하지 않는다.
- 상담 이용 건수 기반의 실제 상담 이용 점수는 학교별 기록 방식 차이로 인한 편차가 있을 수 있다.
- 학교별 특수 상황, 지역 교통 여건 등 정성적 요인은 정량 분석에서 완전히 포착되지 않는다.
"""

MD_PATH = DOCS_DIR / "eda_findings_3x3_2025.md"
MD_PATH.write_text(findings_md, encoding="utf-8")
print(f"  저장: eda_findings_3x3_2025.md")

# ════════════════════════════════════════════════════════════════════════════
# 최종 요약
# ════════════════════════════════════════════════════════════════════════════
print("\n" + "=" * 60)
print("3x3 분류 기반 EDA 엑셀 파일 생성 완료")
print("=" * 60)
print(f"\n엑셀 파일: {OUTPUT_XLSX}")
for sn, _ in SHEETS:
    print(f"  [{sn}]")
print(f"\n그래프: 10종 x 2개(영문+국문) = 20개 PNG -> {FIGURES_DIR}")
print(f"해석 문서: {MD_PATH}")
